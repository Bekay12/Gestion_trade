from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QSpacerItem, QVBoxLayout, QHBoxLayout, QWidget,
    QPushButton, QLabel, QLineEdit, QInputDialog, QListWidget, QListWidgetItem,
    QMessageBox, QProgressDialog, QScrollArea, QSizePolicy, QTableWidget,
    QTableWidgetItem, QComboBox, QHeaderView, QSpinBox, QCheckBox, QTabWidget, QTextEdit
)
from PyQt5.QtWidgets import QAbstractItemView
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QObject, QTimer
from PyQt5.QtGui import QColor
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib import dates as mdates
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
import sys
import os
import traceback
import threading
import faulthandler
from datetime import datetime
import pandas as pd
import yfinance as yf

# Ensure project `src` root is on sys.path
PROJECT_SRC = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if PROJECT_SRC not in sys.path:
    sys.path.insert(0, PROJECT_SRC)

# Segfault mitigation: force Python backtest path in desktop UI unless user overrides.
os.environ.setdefault('QSI_DISABLE_C_ACCELERATION', '1')
# Segfault mitigation: avoid curl_cffi/yfinance recommendation fetches in desktop callbacks.
os.environ.setdefault('QSI_CONSENSUS_OFFLINE', '1')

from qsi import analyse_signaux_populaires, analyse_et_affiche, load_symbols_from_txt, period
from qsi import download_stock_data, backtest_signals, plot_unified_chart, get_trading_signal, resolve_symbol_scoring_context
import qsi
from trading_c_acceleration.qsi_optimized import extract_best_parameters


_CRASH_LOG_FILE = None


def _install_runtime_diagnostics():
    """Installe des hooks pour capturer les crashs/erreurs non gérées dans un fichier."""
    global _CRASH_LOG_FILE
    if _CRASH_LOG_FILE is not None:
        return

    try:
        logs_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'cache_logs'))
        os.makedirs(logs_dir, exist_ok=True)
        log_path = os.path.join(logs_dir, 'desktop_runtime_crash.log')
        _CRASH_LOG_FILE = open(log_path, 'a', encoding='utf-8', buffering=1)
        _CRASH_LOG_FILE.write("\n\n=== Session start: " + datetime.now().isoformat() + " ===\n")
    except Exception:
        _CRASH_LOG_FILE = None
        return

    old_excepthook = sys.excepthook

    def _log_exception(exc_type, exc_value, exc_tb):
        try:
            _CRASH_LOG_FILE.write("\n[Unhandled exception]\n")
            traceback.print_exception(exc_type, exc_value, exc_tb, file=_CRASH_LOG_FILE)
            _CRASH_LOG_FILE.flush()
        except Exception:
            pass
        old_excepthook(exc_type, exc_value, exc_tb)

    sys.excepthook = _log_exception

    if hasattr(threading, 'excepthook'):
        old_thread_excepthook = threading.excepthook

        def _thread_excepthook(args):
            try:
                _CRASH_LOG_FILE.write(f"\n[Unhandled thread exception] thread={getattr(args.thread, 'name', 'unknown')}\n")
                traceback.print_exception(args.exc_type, args.exc_value, args.exc_traceback, file=_CRASH_LOG_FILE)
                _CRASH_LOG_FILE.flush()
            except Exception:
                pass
            old_thread_excepthook(args)

        threading.excepthook = _thread_excepthook

    try:
        faulthandler.enable(file=_CRASH_LOG_FILE, all_threads=True)
    except Exception:
        pass

try:
    from symbol_manager import get_symbols_by_list_type, get_recent_symbols, get_symbol_info_from_db
    SYMBOL_MANAGER_AVAILABLE = True
except ImportError:
    SYMBOL_MANAGER_AVAILABLE = False
    get_symbol_info_from_db = None


def _fetch_yf_info_with_timeout(symbol: str, timeout_sec: float = 2.0) -> dict:
    """Récupère yfinance.info avec timeout court pour éviter les blocages prolongés."""
    executor = ThreadPoolExecutor(max_workers=1)
    future = executor.submit(lambda: yf.Ticker(symbol).info)
    try:
        info = future.result(timeout=timeout_sec)
        return info if isinstance(info, dict) else {}
    except (FuturesTimeoutError, Exception):
        return {}
    finally:
        executor.shutdown(wait=True, cancel_futures=False)


def _is_valid_ticker_info(symbol: str, info: dict) -> bool:
    """Vérifie si un ticker a au moins des données de prix dans yfinance.info."""
    # Cherche des indicateurs qu'il y a des prix réelles (pas de ticker invalide/délisté)
    try:
        # Un ticker valide doit avoir au minimum un prix ou une cap de marché
        has_market_price = info.get('regularMarketPrice') is not None or \
                          info.get('currentPrice') is not None or \
                          info.get('lastPrice') is not None
        has_market_cap = info.get('marketCap') is not None or \
                        info.get('totalAssets') is not None
        has_symbol = info.get('symbol') == symbol
        
        return bool(has_symbol and (has_market_price or has_market_cap))
    except Exception:
        return False


def _get_sector_cache_first(symbol: str) -> str:
    """Récupère le secteur en priorité DB/cache local, puis fallback yfinance timeout court."""
    try:
        if get_symbol_info_from_db:
            db_info = get_symbol_info_from_db(symbol) or {}
            sector = db_info.get('sector')
            if sector and sector != 'Inconnu':
                return sector
    except Exception:
        pass

    try:
        if getattr(qsi, 'get_pickle_cache', None) is not None:
            fin_cache = qsi.get_pickle_cache(symbol, 'financial', ttl_hours=24 * 365)
            if isinstance(fin_cache, dict):
                sector = fin_cache.get('sector')
                if sector and sector != 'Inconnu':
                    return sector
    except Exception:
        pass

    if getattr(qsi, 'OFFLINE_MODE', False):
        return 'Inconnu'

    info = _fetch_yf_info_with_timeout(symbol, timeout_sec=2.0)
    sector = info.get('sector', 'Inconnu') if isinstance(info, dict) else 'Inconnu'
    return sector if sector else 'Inconnu'


class AnalysisThread(QThread):
    """Runs analyse_signaux_populaires in a subprocess for memory isolation."""
    result_ready = pyqtSignal(dict)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, symbols, mes_symbols, period="12mo", analysis_id=0, min_holding_days=7):
        super().__init__()
        self.symbols = symbols
        self.mes_symbols = mes_symbols
        self.period = period
        self._stop_requested = False
        self.analysis_id = analysis_id
        self.min_holding_days = max(1, int(min_holding_days))
        self._process = None

    def run(self):
        import subprocess, pickle, tempfile
        args_file = result_file = None
        try:
            self.progress.emit("Analyse en cours...")
            fiab_threshold = 30
            args = {
                'task': 'analyse_signaux',
                'symbols': self.symbols,
                'mes_symbols': self.mes_symbols,
                'period': self.period,
                'taux_reussite_min': fiab_threshold,
                'min_holding_days': self.min_holding_days,
            }
            fd_a, args_file = tempfile.mkstemp(suffix='_args.pkl')
            fd_r, result_file = tempfile.mkstemp(suffix='_result.pkl')
            os.close(fd_a); os.close(fd_r)
            with open(args_file, 'wb') as f:
                pickle.dump(args, f)

            worker_script = os.path.join(PROJECT_SRC, '_subprocess_worker.py')
            self._process = subprocess.Popen(
                [sys.executable, '-u', worker_script, args_file, result_file],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=PROJECT_SRC,
            )
            for line in self._process.stdout:
                if self._stop_requested:
                    self._process.terminate()
                    break
                print(line, end='', flush=True)
            self._process.wait()

            if self._stop_requested:
                return
            if self._process.returncode != 0:
                self.error.emit(f"Analyse subprocess failed (code {self._process.returncode})")
                return
            with open(result_file, 'rb') as f:
                payload = pickle.load(f)
            if payload.get('success'):
                result = payload['result']
                result['_analysis_id'] = self.analysis_id
                self.result_ready.emit(result)
            else:
                self.error.emit(payload.get('error', 'Unknown subprocess error'))
        except Exception as e:
            if not self._stop_requested:
                self.error.emit(str(e))
        finally:
            for fp in (args_file, result_file):
                if fp:
                    try: os.unlink(fp)
                    except OSError: pass

    def stop(self):
        self._stop_requested = True
        if self._process and self._process.poll() is None:
            self._process.terminate()


class DownloadThread(QThread):
    """Runs download_stock_data in a subprocess for memory isolation."""
    result_ready = pyqtSignal(dict)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, symbols, period="12mo", do_backtest=False, analysis_id=0, min_holding_days=7, best_parameters_snapshot=None):
        super().__init__()
        self.symbols = symbols
        self.period = period
        self._stop_requested = False
        self.analysis_id = analysis_id
        self.min_holding_days = max(1, int(min_holding_days))
        self._process = None

    def run(self):
        import subprocess, pickle, tempfile
        args_file = result_file = None
        try:
            self.progress.emit("Téléchargement des données...")
            args = {
                'task': 'download',
                'symbols': self.symbols,
                'period': self.period,
            }
            fd_a, args_file = tempfile.mkstemp(suffix='_args.pkl')
            fd_r, result_file = tempfile.mkstemp(suffix='_result.pkl')
            os.close(fd_a); os.close(fd_r)
            with open(args_file, 'wb') as f:
                pickle.dump(args, f)

            worker_script = os.path.join(PROJECT_SRC, '_subprocess_worker.py')
            self._process = subprocess.Popen(
                [sys.executable, '-u', worker_script, args_file, result_file],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=PROJECT_SRC,
            )
            for line in self._process.stdout:
                if self._stop_requested:
                    self._process.terminate()
                    break
                print(line, end='', flush=True)
            self._process.wait()

            if self._stop_requested:
                return
            if self._process.returncode != 0:
                self.error.emit(f"Download subprocess failed (code {self._process.returncode})")
                return
            with open(result_file, 'rb') as f:
                payload = pickle.load(f)
            if payload.get('success'):
                result = payload['result']
                result['_analysis_id'] = self.analysis_id
                self.result_ready.emit(result)
            else:
                self.error.emit(payload.get('error', 'Unknown subprocess error'))
        except Exception as e:
            if not self._stop_requested:
                self.error.emit(str(e))
        finally:
            for fp in (args_file, result_file):
                if fp:
                    try: os.unlink(fp)
                    except OSError: pass

    def stop(self):
        self._stop_requested = True
        if self._process and self._process.poll() is None:
            self._process.terminate()

class LogCapture:
    """Captures stdout/stderr and writes both to QTextEdit and to original stdout/stderr.
    Uses a Qt signal to marshal UI writes back to the GUI thread."""
    class _Emitter(QObject):
        message = pyqtSignal(str)

    def __init__(self, text_edit):
        self.text_edit = text_edit
        self.original_stdout = sys.__stdout__
        self.original_stderr = sys.__stderr__
        self._emitter = self._Emitter()
        self._emitter.message.connect(self._append_message)

    def _append_message(self, message):
        """Append only from GUI thread via queued Qt signal delivery."""
        try:
            self.text_edit.append(message)
        except Exception:
            # Widget can be deleted while background logs are still flushing.
            pass
    
    def write(self, message):
        """Write message to QTextEdit and original stdout."""
        try:
            if message and message.strip():
                # Marshal UI updates to Qt main thread.
                self._emitter.message.emit(message.rstrip())
                # Also print to original stdout (terminal)
                self.original_stdout.write(message)
                self.original_stdout.flush()
        except Exception:
            # Fail silently to avoid breaking print calls
            try:
                self.original_stdout.write(message)
                self.original_stdout.flush()
            except Exception:
                pass
    
    def flush(self):
        """Flush the buffer."""
        try:
            self.original_stdout.flush()
        except Exception:
            pass
    
    def isatty(self):
        """Required for some code that checks if stdout is a TTY."""
        return False

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Stock Analysis Tool")
        self.setGeometry(100, 100, 1200, 800)
        self.debug_mode_enabled = False
        self._analysis_running = False
        self._active_analysis_thread = None

        # 🔄 Synchroniser personal et optimization vers popular au démarrage
        if SYMBOL_MANAGER_AVAILABLE:
            try:
                from symbol_manager import sync_all_to_popular
                stats = sync_all_to_popular()
                if stats['total'] > 0:
                    print(f"🔄 Auto-sync au démarrage: {stats['total']} symboles ajoutés à popular")
            except Exception as e:
                print(f"⚠️ Erreur sync auto popular: {e}")

        # Charger les listes au démarrage (SQLite si dispo, sinon txt)
        self.popular_symbols_data = self._load_symbols_preferred("popular_symbols.txt", "popular")
        self.mes_symbols_data = self._load_symbols_preferred("mes_symbols.txt", "personal")
        self.coko_symbols_data = self._load_symbols_preferred("coko_symbols.txt", "coko")
        self.optim_symbols_data = self._load_symbols_preferred("optimisation_symbols.txt", "optimization")
        
        # Tabs-based UI: results-focused navigation
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
        self._charts_dirty = False
        self._charts_refresh_scheduled = False
        self._comparisons_dirty = False
        self._comparisons_refresh_scheduled = False

        # Tab 1: Analyze (input + quick summary)
        self.analyze_container = QWidget()
        self.layout = QVBoxLayout(self.analyze_container)
        self.tabs.addTab(self.analyze_container, "Analyser")

        # Tab 2: Results (detailed table of analysis results)
        self.results_container = QWidget()
        self.results_layout = QVBoxLayout(self.results_container)
        self.tabs.addTab(self.results_container, "Résultats")

        # Tab 3: Charts (per-symbol graphs and metrics)
        self.charts_container = QWidget()
        self.charts_layout = QVBoxLayout(self.charts_container)
        self.charts_scroll = QScrollArea()
        self.charts_scroll.setWidgetResizable(True)
        self.charts_scroll_widget = QWidget()
        self.charts_scroll_layout = QVBoxLayout(self.charts_scroll_widget)
        self.charts_scroll.setWidget(self.charts_scroll_widget)
        self.charts_layout.addWidget(self.charts_scroll)
        self.tabs.addTab(self.charts_container, "Graphiques")

        # Tab 4: Comparisons (multi-symbol visualizations)
        self.comparisons_container = QWidget()
        self.comparisons_layout = QVBoxLayout(self.comparisons_container)
        self.comparisons_layout.addWidget(QLabel("📈 Comparaisons entre symboles (heatmaps, scatter)"))
        self.tabs.addTab(self.comparisons_container, "Comparaisons")

        # Tab 5: Logs (display stdout/stderr)
        self.logs_container = QWidget()
        self.logs_layout = QVBoxLayout(self.logs_container)
        self.logs_text = QTextEdit()
        self.logs_text.setReadOnly(True)
        self.logs_text.setStyleSheet("font-family: monospace; font-size: 9pt;")
        self.logs_layout.addWidget(QLabel("📝 Logs du système"))
        self.logs_layout.addWidget(self.logs_text)
        self.tabs.addTab(self.logs_container, "Logs")
        self.tabs.currentChanged.connect(self._on_tab_changed)
        
        # Setup log capture to redirect stdout/stderr to logs_text AND terminal
        try:
            self.log_capture = LogCapture(self.logs_text)
            sys.stdout = self.log_capture
            sys.stderr = self.log_capture
            print("✅ Log capture activé - les messages s'affichent ici et dans le terminal")
        except Exception as e:
            print(f"⚠️ Erreur initialisation LogCapture: {e}")

        # Build analyze tab UI
        self.setup_ui()

        self.current_results = []
        self._analysis_id = 0  # 🔧 Identifiant unique pour chaque analyse
        self.best_parameters = {}

    def _get_best_parameters_cached(self, force_refresh: bool = False):
        """Return the in-memory best-parameter cache, lazy-loading from SQLite if empty."""
        if not getattr(self, 'best_parameters', None) or force_refresh:
            try:
                self.best_parameters = qsi.extract_best_parameters()
            except Exception:
                self.best_parameters = {}
        return self.best_parameters

    def _on_tab_changed(self, index: int):
        """Refresh heavy tabs only when user actually opens them."""
        try:
            current_widget = self.tabs.widget(index)
            if current_widget is self.charts_container and self._charts_dirty:
                self._schedule_charts_refresh()
            elif current_widget is self.comparisons_container and self._comparisons_dirty:
                self._schedule_comparisons_refresh()
        except Exception:
            pass

    def _schedule_charts_refresh(self):
        """Queue charts refresh on the next GUI loop turn to avoid re-entrancy crashes."""
        if getattr(self, '_charts_refresh_scheduled', False):
            return
        self._charts_refresh_scheduled = True
        QTimer.singleShot(0, self._refresh_charts_tab_safe)

    def _refresh_charts_tab_safe(self):
        self._charts_refresh_scheduled = False
        try:
            # Only render charts when the charts tab is active.
            if self.tabs.currentWidget() is not self.charts_container:
                self._charts_dirty = True
                return
            self.populate_charts_tab()
            self._charts_dirty = False
        except Exception as e:
            print(f"⚠️ Erreur lors de la mise à jour de l'onglet Graphiques: {e}")

    def _schedule_comparisons_refresh(self):
        """Queue comparisons refresh on next GUI loop turn to avoid re-entrancy crashes."""
        if getattr(self, '_comparisons_refresh_scheduled', False):
            return
        self._comparisons_refresh_scheduled = True
        QTimer.singleShot(0, self._refresh_comparisons_tab_safe)

    def _refresh_comparisons_tab_safe(self):
        self._comparisons_refresh_scheduled = False
        try:
            if self.tabs.currentWidget() is not self.comparisons_container:
                self._comparisons_dirty = True
                return
            self.populate_comparisons_tab()
            self._comparisons_dirty = False
        except Exception as e:
            print(f"⚠️ Erreur lors de la mise à jour de l'onglet Comparaisons: {e}")

    def _schedule_result_visuals_refresh(self, result, mode: str):
        """Queue heavy chart rendering after the current event loop turn."""
        self._pending_visuals_result = result
        self._pending_visuals_mode = mode
        if getattr(self, '_result_visuals_refresh_scheduled', False):
            return
        self._result_visuals_refresh_scheduled = True
        QTimer.singleShot(0, self._refresh_result_visuals_safe)

    def _refresh_result_visuals_safe(self):
        self._result_visuals_refresh_scheduled = False
        result = getattr(self, '_pending_visuals_result', None)
        mode = getattr(self, '_pending_visuals_mode', '')
        if result is None:
            return

        try:
            if mode == 'download':
                self._render_download_result_visuals(result)
            else:
                self._render_analysis_result_visuals(result)
            # Domain charts tab must be recomputed after results visuals update.
            self._charts_dirty = True
            if self.tabs.currentWidget() is self.charts_container:
                self._schedule_charts_refresh()
        except Exception as e:
            print(f"⚠️ Erreur lors du rendu différé des graphiques: {e}")

    def _render_download_result_visuals(self, result):
        """Render the downloadable analysis charts outside the completion callback."""
        try:
            self.clear_plots()

            if not isinstance(result, dict):
                return

            filtered = getattr(self, 'current_results', []) or []
            data = result.get('data', {}) or {}

            # Keep only symbols that passed the current fiabilité filter.
            min_val = self.fiab_threshold_spin.value() if hasattr(self, 'fiab_threshold_spin') else 30
            include_none_val = True
            filtered_symbols = []
            for r in filtered:
                fiab = r.get('Fiabilite', 'N/A')
                nb_trades = r.get('NbTrades', 0)
                try:
                    if int(nb_trades) > 0 and int(nb_trades) < min_val and not include_none_val:
                        continue
                except Exception:
                    if not include_none_val:
                        continue
                try:
                    if fiab == 'N/A':
                        if include_none_val:
                            filtered_symbols.append(r.get('Symbole'))
                    elif float(fiab) >= float(min_val):
                        filtered_symbols.append(r.get('Symbole'))
                except Exception:
                    if include_none_val:
                        filtered_symbols.append(r.get('Symbole'))

            rendered_count = 0
            for sym in [s for s in filtered_symbols if s]:
                stock_data = data.get(sym)
                if not stock_data:
                    continue
                prices = stock_data['Close']
                volumes = stock_data['Volume']
                row = next((r for r in filtered if r.get('Symbole') == sym), {})
                precomp = {
                    'signal': row.get('Signal'),
                    'last_price': row.get('Prix'),
                    'trend': row.get('Tendance'),
                    'last_rsi': row.get('RSI'),
                    'volume_moyen': row.get('Volume moyen'),
                    'score': row.get('Score'),
                    'domaine': row.get('Domaine'),
                    'cap_range': row.get('CapRange'),
                }
                fig = self._build_symbol_figure_with_score(sym, prices, volumes, precomp=precomp, events=[])
                canvas = FigureCanvas(fig)
                canvas.setMinimumHeight(520)
                self.plots_layout.addWidget(canvas)
                rendered_count += 1

            if filtered_symbols and rendered_count == 0:
                self.plots_layout.addWidget(QLabel("Aucun graphe integre n'a pu etre affiche pour cette analyse."))
        except Exception:
            pass

    def _render_analysis_result_visuals(self, result):
        """Render the final backtest charts outside the completion callback."""
        try:
            self.clear_plots()

            top_buys = result.get('top_achats_fiables', []) if isinstance(result, dict) else []
            top_sells = result.get('top_ventes_fiables', []) if isinstance(result, dict) else []
            backtests = result.get('backtest_results', []) if isinstance(result, dict) else []
            events_map = {bt.get('Symbole'): bt.get('events', []) for bt in backtests}
            score_series_map = {
                bt.get('Symbole'): {
                    'score_dates': bt.get('score_dates', []),
                    'score_values': bt.get('score_values', []),
                    'seuil_achat': bt.get('seuil_achat'),
                    'seuil_vente': bt.get('seuil_vente'),
                }
                for bt in backtests
            }
            existing_data = result.get('data', {}) if isinstance(result, dict) else {}

            def _get_stock_data_for_symbol(sym):
                stock_data = existing_data.get(sym)
                if stock_data:
                    return stock_data
                # Fallback minimal when analysis payload does not include data.
                try:
                    return download_stock_data([sym], period=self.period_input.currentData() or '15mo').get(sym)
                except Exception:
                    return None

            def embed_symbol_list(symbol_list):
                if not symbol_list:
                    return
                for s in symbol_list:
                    sym = s['Symbole'] if isinstance(s, dict) and 'Symbole' in s else s
                    try:
                        stock_data = _get_stock_data_for_symbol(sym)
                        if not stock_data:
                            continue
                        prices = stock_data['Close']
                        volumes = stock_data['Volume']
                        pre_row = next((r for r in self.current_results if r.get('Symbole') == sym), s if isinstance(s, dict) else {})
                        precomp = {
                            'signal': pre_row.get('Signal'),
                            'last_price': pre_row.get('Prix'),
                            'trend': pre_row.get('Tendance'),
                            'last_rsi': pre_row.get('RSI'),
                            'volume_moyen': pre_row.get('Volume moyen'),
                            'score': pre_row.get('Score'),
                            'domaine': pre_row.get('Domaine'),
                            'cap_range': pre_row.get('CapRange'),
                            'score_dates': score_series_map.get(sym, {}).get('score_dates', []),
                            'score_values': score_series_map.get(sym, {}).get('score_values', []),
                            'seuil_achat': score_series_map.get(sym, {}).get('seuil_achat'),
                            'seuil_vente': score_series_map.get(sym, {}).get('seuil_vente'),
                        }
                        events = events_map.get(sym, [])
                        if len(events) == 0:
                            print(f"⚠️ {sym}: Aucun événement généré")
                        else:
                            print(f"✅ {sym}: {len(events)} événement(s) trouvé(s)")
                        fig = self._build_symbol_figure_with_score(sym, prices, volumes, precomp=precomp, events=events)
                        canvas = FigureCanvas(fig)
                        canvas.setMinimumHeight(520)
                        self.plots_layout.addWidget(canvas)
                    except Exception:
                        continue

            embed_symbol_list(top_buys)
            embed_symbol_list(top_sells)
        except Exception:
            pass
    
    def _debug_log(self, message: str):
        """Affiche un log uniquement si le mode debug est actif."""
        if self.debug_mode_enabled:
            print(message)

    def _load_symbols_preferred(self, filename: str, list_type: str):
        """Charge depuis SQLite si possible, sinon depuis le fichier txt."""
        symbols = []
        if SYMBOL_MANAGER_AVAILABLE:
            try:
                symbols = get_symbols_by_list_type(list_type, active_only=True)
                print(f"✅ {list_type}: {len(symbols)} symboles chargés depuis SQLite")
            except Exception as e:
                print(f"⚠️ Erreur chargement SQLite pour {list_type}: {e}")
                symbols = []
        if not symbols:
            try:
                symbols = load_symbols_from_txt(filename, use_sqlite=True)
                print(f"✅ {list_type}: {len(symbols)} symboles chargés depuis fichier/SQLite")
            except Exception as e:
                print(f"⚠️ Erreur chargement {filename}: {e}")
                symbols = []
        # Dédupe en conservant l'ordre
        return list(dict.fromkeys([s for s in symbols if s]))
    def setup_ui(self):
        # Title
        # title_label = QLabel("Stock Analysis Tool")
        # title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        # self.layout.addWidget(title_label)

        # Input de symbole
        self.symbol_input = QLineEdit()
        self.symbol_input.setPlaceholderText("Enter stock symbol (e.g., AAPL)")
        self.layout.addWidget(self.symbol_input)

        # Listes de symboles
        lists_container = QHBoxLayout()
        lists_container.setSpacing(24)  # Ajuste ce chiffre, ex: 24px entre les trois zones

        # Liste populaire
        popular_sorted = sorted(self.popular_symbols_data)
        popular_layout = QHBoxLayout()
        popular_listcol = QVBoxLayout()
        self.popular_label = QLabel()
        self.popular_label.setAlignment(Qt.AlignCenter)
        self.popular_label.setWordWrap(True)
        popular_layout.addWidget(self.popular_label)
        self.popular_list = QListWidget()
        self.popular_list.setMaximumHeight(70)
        for s in popular_sorted:
            if s:
                item = QListWidgetItem(s)
                item.setData(Qt.UserRole, s)
                self.popular_list.addItem(item)
        self.popular_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        popular_listcol.addWidget(self.popular_list)
        popular_layout.addLayout(popular_listcol)

        pop_btns = QVBoxLayout()
        pop_btns.setSpacing(2)
        self.pop_add_btn = QPushButton("Ajouter")
        self.pop_del_btn = QPushButton("Supprimer")
        self.pop_show_btn = QPushButton("Afficher")
        pop_btns.addWidget(self.pop_add_btn)
        pop_btns.addWidget(self.pop_del_btn)
        pop_btns.addWidget(self.pop_show_btn)
        popular_layout.addLayout(pop_btns)

        lists_container.addLayout(popular_layout)
       
        lists_container.addItem(QSpacerItem(48, 20, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))  # espace “élastique” mais raisonnable

        # Liste personnelle
        mes_sorted = sorted(self.mes_symbols_data)
        mes_layout = QHBoxLayout()
        mes_listcol = QVBoxLayout()
        self.mes_label = QLabel()
        self.mes_label.setAlignment(Qt.AlignCenter)
        self.mes_label.setWordWrap(True)
        mes_layout.addWidget(self.mes_label)
        self.mes_list = QListWidget()
        self.mes_list.setMaximumHeight(70)
        for s in mes_sorted:
            if s:
                item = QListWidgetItem(s)
                item.setData(Qt.UserRole, s)
                self.mes_list.addItem(item)
        self.mes_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        mes_listcol.addWidget(self.mes_list)
        mes_layout.addLayout(mes_listcol)

        mes_btns = QVBoxLayout()
        mes_btns.setSpacing(2)
        self.mes_add_btn = QPushButton("Ajouter")
        self.mes_del_btn = QPushButton("Supprimer")
        self.mes_show_btn = QPushButton("Afficher")
        mes_btns.addWidget(self.mes_add_btn)
        mes_btns.addWidget(self.mes_del_btn)
        mes_btns.addWidget(self.mes_show_btn)
        mes_layout.addLayout(mes_btns)

        lists_container.addLayout(mes_layout)

        lists_container.addItem(QSpacerItem(48, 20, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))

        # Liste coko
        coko_sorted = sorted(self.coko_symbols_data)
        coko_layout = QHBoxLayout()
        coko_listcol = QVBoxLayout()
        self.coko_label = QLabel()
        self.coko_label.setAlignment(Qt.AlignCenter)
        self.coko_label.setWordWrap(True)
        coko_layout.addWidget(self.coko_label)
        self.coko_list = QListWidget()
        self.coko_list.setMaximumHeight(70)
        for s in coko_sorted:
            if s:
                item = QListWidgetItem(s)
                item.setData(Qt.UserRole, s)
                self.coko_list.addItem(item)
        self.coko_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        coko_listcol.addWidget(self.coko_list)
        coko_layout.addLayout(coko_listcol)

        coko_btns = QVBoxLayout()
        coko_btns.setSpacing(2)
        self.coko_add_btn = QPushButton("Ajouter")
        self.coko_del_btn = QPushButton("Supprimer")
        self.coko_show_btn = QPushButton("Afficher")
        coko_btns.addWidget(self.coko_add_btn)
        coko_btns.addWidget(self.coko_del_btn)
        coko_btns.addWidget(self.coko_show_btn)
        coko_layout.addLayout(coko_btns)

        lists_container.addLayout(coko_layout)

        lists_container.addItem(QSpacerItem(48, 20, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))

        # ========== NOUVELLES LISTES OPTIMISATION ==========
        # Liste 1 : 30 symboles ALÉATOIRES
        random_layout = QHBoxLayout()
        random_listcol = QVBoxLayout()
        self.random_label = QLabel()
        self.random_label.setAlignment(Qt.AlignCenter)
        self.random_label.setWordWrap(True)
        random_layout.addWidget(self.random_label)
        self.random_list = QListWidget()
        self.random_list.setMaximumHeight(70)
        self.random_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        random_listcol.addWidget(self.random_list)
        random_layout.addLayout(random_listcol)
        
        random_btns = QVBoxLayout()
        random_btns.setSpacing(2)
        self.random_refresh_btn = QPushButton("🔄 Nouveau")
        self.random_refresh_btn.clicked.connect(self.refresh_random_symbols)
        self.random_show_btn = QPushButton("Afficher")
        self.random_all_btn = QPushButton("📋 Tout sélect.")
        self.random_all_btn.clicked.connect(lambda: self._select_all_items(self.random_list))
        random_btns.addWidget(self.random_refresh_btn)
        random_btns.addWidget(self.random_show_btn)
        random_btns.addWidget(self.random_all_btn)
        random_layout.addLayout(random_btns)
        
        lists_container.addLayout(random_layout)
        
        lists_container.addItem(QSpacerItem(48, 5, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))
        
        # Liste 2 : 30 derniers SYMBOLES AJOUTÉS
        recent_layout = QHBoxLayout()
        recent_listcol = QVBoxLayout()
        self.recent_label = QLabel()
        self.recent_label.setAlignment(Qt.AlignCenter)
        self.recent_label.setWordWrap(True)
        recent_layout.addWidget(self.recent_label)
        self.recent_list = QListWidget()
        self.recent_list.setMaximumHeight(70)
        self.recent_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        recent_listcol.addWidget(self.recent_list)
        recent_layout.addLayout(recent_listcol)
        
        recent_btns = QVBoxLayout()
        recent_btns.setSpacing(2)
        self.recent_show_btn = QPushButton("Afficher")
        self.recent_all_btn = QPushButton("📋 Tout sélect.")
        self.recent_all_btn.clicked.connect(lambda: self._select_all_items(self.recent_list))
        recent_btns.addWidget(self.recent_show_btn)
        recent_btns.addWidget(self.recent_all_btn)
        recent_layout.addLayout(recent_btns)
        
        lists_container.addLayout(recent_layout)

        # Raccourcis top movers placés à l'extrême droite de la ligne des listes
        lists_container.addItem(QSpacerItem(80, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        movers_layout = QVBoxLayout()
        movers_layout.setSpacing(6)
        movers_layout.addWidget(QLabel("Top Movers (Jour)"))
        self.top_winners_btn = QPushButton("Top 30 Winners")
        self.top_winners_btn.setToolTip("Charge les 30 plus fortes hausses du jour")
        self.top_winners_btn.clicked.connect(lambda: self._show_top_daily_movers('winners'))
        movers_layout.addWidget(self.top_winners_btn)

        self.top_losers_btn = QPushButton("Top 30 Losers")
        self.top_losers_btn.setToolTip("Charge les 30 plus fortes baisses du jour")
        self.top_losers_btn.clicked.connect(lambda: self._show_top_daily_movers('losers'))
        movers_layout.addWidget(self.top_losers_btn)
        lists_container.addLayout(movers_layout)

        self.layout.addLayout(lists_container)

        top_controls = QHBoxLayout()

        # Période d'analyse à gauche (menu déroulant)
        top_controls.addWidget(QLabel("Période d'analyse:"))
        self.period_input = QComboBox()
        self.period_input.setMinimumWidth(220)
        period_options = [
            ("3mo",  "3 mois   — ~63 points (journalier)"),
            ("6mo",  "6 mois   — ~126 points (journalier)"),
            ("1y",   "1 an     — ~252 points (journalier)"),
            ("15mo", "15 mois  — ~315 points (journalier)"),
            ("18mo", "18 mois  — ~378 points (journalier)"),
            ("2y",   "2 ans    — ~504 points (journalier)"),
            ("3y",   "3 ans    — ~756 points (journalier)"),
            ("4y",   "4 ans    — ~1 008 points (journalier)"),
            ("5y",   "5 ans    — ~1 260 points (journalier)"),
            ("10y",  "10 ans   — ~2 520 points (journalier)"),
            ("max",  "Max      — historique complet (journalier)"),
        ]
        default_index = 0
        for i, (value, label) in enumerate(period_options):
            self.period_input.addItem(label, userData=value)
            if value == period:
                default_index = i
        self.period_input.setCurrentIndex(default_index)
        top_controls.addWidget(self.period_input)

        top_controls.addSpacing(24)  # Petit espace pour l'esthétique

        # Boutons d'analyse sur la même ligne
        self.analyze_button = QPushButton("Analyze")
        self.analyze_button.clicked.connect(self.analyze_stock)
        top_controls.addWidget(self.analyze_button)

        self.backtest_button = QPushButton("Analyze and Backtest")
        self.backtest_button.clicked.connect(self.analyse_and_backtest)
        top_controls.addWidget(self.backtest_button)

        # Seuil minimum de fiabilité pour le backtest (à droite du bouton backtest)
        top_controls.addWidget(QLabel("Seuil fiabilité:"))
        self.fiab_threshold_spin = QSpinBox()
        self.fiab_threshold_spin.setMinimum(0)
        self.fiab_threshold_spin.setMaximum(100)
        self.fiab_threshold_spin.setValue(30)
        self.fiab_threshold_spin.setSuffix("%")
        self.fiab_threshold_spin.setMaximumWidth(80)
        self.fiab_threshold_spin.setToolTip("Seuil minimum de fiabilité pour filtrer les résultats du backtest")
        top_controls.addWidget(self.fiab_threshold_spin)

        # Durée minimale de détention (en jours de bourse / barres actives)
        top_controls.addWidget(QLabel("Durée min position:"))
        self.min_hold_days_spin = QSpinBox()
        self.min_hold_days_spin.setMinimum(1)
        self.min_hold_days_spin.setMaximum(60)
        self.min_hold_days_spin.setValue(7)
        self.min_hold_days_spin.setSuffix(" j")
        self.min_hold_days_spin.setMaximumWidth(80)
        self.min_hold_days_spin.setToolTip("Nombre minimum de jours actifs avant d'autoriser une vente en backtest")
        top_controls.addWidget(self.min_hold_days_spin)

        top_controls.addSpacing(24)  # Petit espace pour l'esthétique

        self.toggle_bottom_btn = QPushButton("Masquer détails")
        self.toggle_bottom_btn.setCheckable(True)
        self.toggle_bottom_btn.clicked.connect(self.toggle_bottom)
        top_controls.addWidget(self.toggle_bottom_btn)
        
        # Bouton pour basculer entre mode online/offline
        self.offline_mode_btn = QPushButton("🌐 Mode: ONLINE")
        self.offline_mode_btn.setCheckable(True)
        self.offline_mode_btn.setChecked(False)
        self.offline_mode_btn.clicked.connect(self.toggle_offline_mode)
        self.offline_mode_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
        top_controls.addWidget(self.offline_mode_btn)

        # Bouton mode debug (désactive les logs en boucle par défaut)
        self.debug_mode_btn = QPushButton("🐞 Debug: OFF")
        self.debug_mode_btn.setCheckable(True)
        self.debug_mode_btn.setChecked(False)
        self.debug_mode_btn.clicked.connect(self.toggle_debug_mode)
        self.debug_mode_btn.setStyleSheet("QPushButton { background-color: #9E9E9E; color: white; font-weight: bold; }")
        self.debug_mode_btn.setToolTip("Active les logs détaillés (secteur, seuils, diagnostics)")
        top_controls.addWidget(self.debug_mode_btn)
        
        # 💾 Bouton pour sauvegarder les graphiques en PDF
        self.save_pdf_btn = QPushButton("💾 Sauvegarder (PDF)")
        self.save_pdf_btn.setToolTip("Sauvegarder tous les graphiques de l'analyse en PDF")
        self.save_pdf_btn.clicked.connect(self.export_results_pdf)
        top_controls.addWidget(self.save_pdf_btn)

        self.layout.addLayout(top_controls)

        # Plots area
        self.plots_scroll = QScrollArea()
        self.plots_scroll.setWidgetResizable(True)
        self.plots_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.plots_container = QWidget()
        self.plots_layout = QVBoxLayout(self.plots_container)
        self.plots_scroll.setWidget(self.plots_container)
        self.plots_scroll.setMinimumHeight(390)

        # Use a vertical splitter so plots remain visible
        from PyQt5.QtWidgets import QSplitter, QTextEdit
        self.splitter = QSplitter(Qt.Vertical)
        self.splitter.addWidget(self.plots_scroll)

        # Bottom container for Analyze tab summary
        bottom_container = QWidget()
        bottom_layout = QVBoxLayout(bottom_container)

        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setMinimumHeight(80)
        self.summary_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        bottom_layout.addWidget(self.summary_text)
        
        # Single merged table combining signals + backtest metrics (will be shown in Results tab)
        self.merged_table = QTableWidget()
        self.merged_table.setMinimumHeight(600)
        self.merged_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        merged_columns = [
        'Symbole','Signal','Score','Prix\n(USD)','Tendance','RSI','Volume\nmoyen($)','Domaine','Cap\nRange','Score/\nSeuil',
        'Fiabilite\n(%)','Nb\nTrades','Gagnants',
        # COLONNES FINANCIÈRES
        'Rev.\nGrowth(%)','EBITDA\nYield(%)','FCF\nYield(%)','D/E\nRatio','Market\nCap(B$)','ROE\n(%)',
        # COLONNES DERIVÉES
        'dPrice','Var5j\n(%)','dRSI','dVol\nRel',
        # COLONNES BACKTEST
        'Gain\ntotal($)','Gain\nmoyen($)',
        # INFO
        'Consensus'
        ]
        # Add table to Results tab, not Analyze tab
        # 🔧 Boutons d'export dans l'onglet Résultats
        export_buttons_layout = QHBoxLayout()
        export_buttons_layout.addStretch()
        self.export_csv_btn = QPushButton("📥 Exporter (CSV)")
        self.export_csv_btn.setToolTip("Exporter les résultats en fichier CSV")
        self.export_csv_btn.clicked.connect(self.export_results_csv)
        self.export_excel_btn = QPushButton("📊 Exporter (Excel)")
        self.export_excel_btn.setToolTip("Exporter les résultats en fichier Excel")
        self.export_excel_btn.clicked.connect(self.export_results_excel)
        export_buttons_layout.addWidget(self.export_csv_btn)
        export_buttons_layout.addWidget(self.export_excel_btn)
        
        self.results_layout.addWidget(QLabel("📋 Résultats détaillés de l'analyse"))
        self.results_layout.addLayout(export_buttons_layout)
        self.results_layout.addWidget(self.merged_table)

        self.merged_table.setColumnCount(len(merged_columns))
        self.merged_table.setHorizontalHeaderLabels(merged_columns)
        self.merged_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # Réduire la hauteur des en-têtes pour 2 lignes maximum
        self.merged_table.horizontalHeader().setMinimumHeight(40)
        # Centrer les en-têtes horizontalement et verticalement
        header_style = """
            QHeaderView::section {
                padding: 4px;
                text-align: center;
                background-color: #f0f0f0;
            }
        """
        self.merged_table.horizontalHeader().setStyleSheet(header_style)
        # Allow sorting by clicking headers (we also provide numeric data via Qt.EditRole)
        self.merged_table.setSortingEnabled(True)
        
        # Keep bottom_container for Analyze tab (only summary_text)
        # NOTA: self.merged_table is now in Results tab, not in Analyze tab
        bottom_layout.addWidget(self.summary_text)  # Only summary in Analyze tab
        self.bottom_container = bottom_container

        self.splitter.addWidget(bottom_container)
        self.splitter.setStretchFactor(0, 3)
        self.splitter.setStretchFactor(1, 1)
        # Try to set an initial sensible ratio so plots are visible by default
        try:
            total_h = max(600, self.height())
            top_h = int(total_h * 0.72)
            bottom_h = total_h - top_h
            self.splitter.setSizes([top_h, bottom_h])
        except Exception:
            pass

        self.layout.addWidget(self.splitter)

        # Connexions des boutons
        self.pop_add_btn.clicked.connect(lambda: self.add_symbol(self.popular_list, "popular_symbols.txt"))
        self.pop_del_btn.clicked.connect(lambda: self.remove_selected(self.popular_list, "popular_symbols.txt"))
        self.pop_show_btn.clicked.connect(lambda: self.show_selected(self.popular_list))
        self.mes_add_btn.clicked.connect(lambda: self.add_symbol(self.mes_list, "mes_symbols.txt"))
        self.mes_del_btn.clicked.connect(lambda: self.remove_selected(self.mes_list, "mes_symbols.txt"))
        self.mes_show_btn.clicked.connect(lambda: self.show_selected(self.mes_list))
        self.coko_add_btn.clicked.connect(lambda: self.add_symbol(self.coko_list, "coko_symbols.txt"))
        self.coko_del_btn.clicked.connect(lambda: self.remove_selected(self.coko_list, "coko_symbols.txt"))
        self.coko_show_btn.clicked.connect(lambda: self.show_selected(self.coko_list))
        
        # Callbacks pour les nouvelles listes
        self.random_show_btn.clicked.connect(lambda: self.show_selected(self.random_list))
        self.recent_show_btn.clicked.connect(lambda: self.show_selected(self.recent_list))
        
        self._update_list_counts()
        
        # Charger les listes aléatoires et récentes au démarrage
        self.refresh_random_symbols()
        self.load_recent_symbols()
    
    def validate_ticker(self, symbol):
        """Validation: ticker valide s'il existe dans yfinance avec prix/marketcap."""
        try:
            # En mode offline, on se limite aux infos locales
            if getattr(qsi, 'OFFLINE_MODE', False):
                sector = _get_sector_cache_first(symbol)
                return bool(sector and sector != 'Inconnu')

            # Récupère les infos yfinance (une seule fois)
            info = _fetch_yf_info_with_timeout(symbol, timeout_sec=2.0)
            
            # Valide que le ticker est réel et a au moins des prix ou une marketcap
            is_valid = _is_valid_ticker_info(symbol, info)
            
            # Si yfinance timeout, fallback au cache/DB
            if not is_valid:
                sector = _get_sector_cache_first(symbol)
                is_valid = bool(sector and sector != 'Inconnu')
            
            return is_valid
        except Exception:
            return False


    def _map_list_type(self, filename: str) -> str:
        lower = filename.lower()
        if 'mes_symbol' in lower:
            return 'personal'
        if 'coko_symbol' in lower:
            return 'coko'
        if 'optimisation' in lower or 'optimization' in lower:
            return 'optimization'
        return 'popular'

    def add_symbol(self, list_widget, filename):
        """Ajoute un ou plusieurs symboles (séparés par des virgules) à la liste.
        Les symboles sont validés, ajoutés individuellement, et la liste est 
        triée alphabétiquement. Si c'est mes_symbols, ils sont aussi ajoutés 
        automatiquement aux symboles populaires.
        """
        text, ok = QInputDialog.getText(
            self, 
            "Ajouter symbole(s)", 
            "Symbole(s) (ex: AAPL ou AAPL, MSFT, GOOGL):"
        )
        
        if ok and text:
            # Identifier si c'est la liste mes_symbols
            is_mes_list = (list_widget == self.mes_list)
            is_coko_list = hasattr(self, 'coko_list') and (list_widget == self.coko_list)
            main_list = list_widget
            secondary_list = self.popular_list if (is_mes_list or is_coko_list) else None
            
            # Parser les symboles séparés par des virgules
            symbols = [s.strip().upper() for s in text.split(",") if s.strip()]
            
            if not symbols:
                return
            
            # Ajouter chaque symbole individuellement
            added_symbols = []
            
            for symbol in symbols:
                if not symbol:
                    continue
                
                # Vérifier que le symbole n'existe pas déjà
                exists_main = any(
                    main_list.item(i).text() == symbol 
                    for i in range(main_list.count())
                )
                
                if exists_main:
                    QMessageBox.information(
                        self, 
                        "Info", 
                        f"{symbol} existe déjà dans la liste principale"
                    )
                    continue
                
                # Validation du ticker
                progress = QProgressDialog(
                    f"Validation de {symbol}...", 
                    None, 0, 0, self
                )
                progress.setWindowModality(Qt.WindowModal)
                progress.setMinimumDuration(0)
                progress.show()
                QApplication.processEvents()
                
                is_valid = self.validate_ticker(symbol)
                progress.close()
                
                if not is_valid:
                    QMessageBox.warning(
                        self,
                        "Ticker invalide",
                        f"Le symbole '{symbol}' n'est pas valide.\\n"
                        "Vérifiez l'orthographe ou consultez Yahoo Finance."
                    )
                    continue
                
                # Ajouter à la liste principale
                item = QListWidgetItem(symbol)
                item.setData(Qt.UserRole, symbol)
                main_list.addItem(item)
                added_symbols.append(symbol)
                
                # Si c'est mes_symbols, ajouter aussi automatiquement aux populaires
                if (is_mes_list or is_coko_list) and secondary_list:
                    exists_secondary = any(
                        secondary_list.item(i).text() == symbol 
                        for i in range(secondary_list.count())
                    )
                    
                    if not exists_secondary:
                        item_pop = QListWidgetItem(symbol)
                        item_pop.setData(Qt.UserRole, symbol)
                        secondary_list.addItem(item_pop)
            
            # Trier les deux listes alphabétiquement
            self._sort_list_alphabetically(main_list)
            if secondary_list:
                self._sort_list_alphabetically(secondary_list)
            
            # Sauvegarder les listes triées
            try:
                from qsi import save_symbols_to_txt
                
                symbols_main = [
                    main_list.item(i).data(Qt.UserRole) 
                    if main_list.item(i).data(Qt.UserRole) is not None 
                    else main_list.item(i).text() 
                    for i in range(main_list.count())
                ]
                save_symbols_to_txt(symbols_main, filename)
                
                # 🔧 Synchroniser avec SQLite après sauvegarde txt
                if SYMBOL_MANAGER_AVAILABLE:
                    try:
                        from symbol_manager import sync_txt_to_sqlite
                        list_type = self._map_list_type(filename)
                        sync_txt_to_sqlite(filename, list_type=list_type)
                        print(f"✅ SQLite synchronisé pour {filename}")
                    except Exception as e:
                        print(f"⚠️ Erreur lors de la sync SQLite: {e}")
                
                if secondary_list:
                    filename_secondary = "popular_symbols.txt"
                    symbols_secondary = [
                        secondary_list.item(i).data(Qt.UserRole) 
                        if secondary_list.item(i).data(Qt.UserRole) is not None 
                        else secondary_list.item(i).text() 
                        for i in range(secondary_list.count())
                    ]
                    save_symbols_to_txt(symbols_secondary, filename_secondary)
                    
                    # 🔧 Synchroniser la liste secondaire avec SQLite
                    if SYMBOL_MANAGER_AVAILABLE:
                        try:
                            from symbol_manager import sync_txt_to_sqlite
                            sync_txt_to_sqlite(filename_secondary, list_type='popular')
                            print(f"✅ SQLite synchronisé pour {filename_secondary}")
                        except Exception as e:
                            print(f"⚠️ Erreur lors de la sync SQLite: {e}")

                # Rafraîchir les compteurs après ajouts/sauvegardes
                self._update_list_counts()
            
            except Exception:
                pass

    def _sort_list_alphabetically(self, list_widget):
        """Trie les éléments d'une QListWidget alphabétiquement."""
        items = []
        
        # Récupérer tous les éléments
        for i in range(list_widget.count()):
            item = list_widget.item(i)
            text = item.text()
            data = item.data(Qt.UserRole)
            items.append((text, data))
        
        # Trier alphabétiquement
        items.sort(key=lambda x: x)
        
        # Vider la liste
        list_widget.clear()
        
        # Réajouter les éléments triés
        for text, data in items:
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, data)
            list_widget.addItem(item)

        # Mettre à jour les compteurs après réinjection
        self._update_list_counts()

    def _update_list_counts(self):
        """Met à jour les libellés avec le nombre d'éléments de chaque liste."""
        try:
            pop_count = self.popular_list.count() if hasattr(self, "popular_list") else 0
            mes_count = self.mes_list.count() if hasattr(self, "mes_list") else 0
            coko_count = self.coko_list.count() if hasattr(self, "coko_list") else 0
            optim_count = self.optim_list.count() if hasattr(self, "optim_list") else 0
            if hasattr(self, "popular_label"):
                self.popular_label.setText(f"Symboles\npopulaires ({pop_count})")
            if hasattr(self, "mes_label"):
                self.mes_label.setText(f"Mes\nsymboles ({mes_count})")
            if hasattr(self, "coko_label"):
                self.coko_label.setText(f"Symboles\ncoko ({coko_count})")
            if hasattr(self, "optim_label"):
                self.optim_label.setText(f"Symboles\noptimisation ({optim_count})")
        except Exception:
            pass


    def remove_selected(self, list_widget, filename):
        items = list_widget.selectedItems()
        if not items:
            QMessageBox.information(self, "Info", "Veuillez sélectionner au moins un symbole à supprimer")
            return
        for it in items:
            list_widget.takeItem(list_widget.row(it))
        self._update_list_counts()
        try:
            from qsi import save_symbols_to_txt
            symbols = [list_widget.item(i).data(Qt.UserRole) if list_widget.item(i).data(Qt.UserRole) is not None else list_widget.item(i).text() for i in range(list_widget.count())]
            save_symbols_to_txt(symbols, filename)
            
            # 🔧 Synchroniser avec SQLite après suppression
            if SYMBOL_MANAGER_AVAILABLE:
                try:
                    from symbol_manager import sync_txt_to_sqlite
                    # Déterminer le type de liste pour SQLite
                    list_type = self._map_list_type(filename)
                    sync_txt_to_sqlite(filename, list_type=list_type)
                    print(f"✅ SQLite synchronisé (suppression) pour {filename}")
                except Exception as e:
                    print(f"⚠️ Erreur lors de la sync SQLite: {e}")
        except Exception:
            pass
        self._update_list_counts()

    def show_selected(self, list_widget):
        items = list_widget.selectedItems()
        if not items:
            QMessageBox.information(self, "Info", "Veuillez sélectionner au moins un symbole à afficher")
            return
        symbols = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in items]
        self.symbol_input.setText(", ".join(symbols))

    def on_download_complete(self, result):
        # 🔧 Vérifier que ce résultat appartient à l'analyse actuelle
        received_id = result.get('_analysis_id', 0) if isinstance(result, dict) else 0
        if received_id != self._analysis_id:
            print(f"⚠️ Résultat ignoré: ID={received_id}, ID actuel={self._analysis_id}")
            return

        self._analysis_running = False
        self._active_analysis_thread = None
        
        # Called when the DownloadThread finishes
        # Re-enable buttons
        self.analyze_button.setEnabled(True)
        self.backtest_button.setEnabled(True)

        if self.progress:
            self.progress.close()
        
        # 🔧 Supprimer filtered_results si existant pour forcer l'utilisation de current_results
        if hasattr(self, 'filtered_results'):
            delattr(self, 'filtered_results')

        data = result.get('data', {}) if isinstance(result, dict) else {}
        backtests = result.get('backtest_results', []) if isinstance(result, dict) else []
        best_params_all = self._get_best_parameters_cached()
        # Build result rows (collect data first, then filter & render plots only for filtered symbols)
        self.current_results = []

        for symbol, stock_data in data.items():
            try:
                prices = stock_data['Close']
                volumes = stock_data['Volume']

                # ✅ Résolution unifiée via resolve_symbol_scoring_context
                # (même logique que process_symbol et _compute_score_series)
                sig = "NEUTRE"
                last_price = float(prices.iloc[-1]) if len(prices) > 0 else 0.0
                trend = False
                last_rsi = 0.0
                volume_mean = float(volumes.mean()) if len(volumes) > 0 else 0.0
                score = 0.0
                derivatives = {}

                score_context = qsi.resolve_symbol_scoring_context(
                    symbol,
                    best_params=best_params_all,
                )
                domaine = score_context['domaine']
                cap_range = score_context['cap_range']
                original_domaine = domaine
                seuil_achat_opt = score_context['seuil_achat']
                seuil_vente_opt = score_context['seuil_vente']
                extras_to_use = score_context['price_extras']
                self._debug_log(f"🔍 {symbol}: ctx domaine={domaine} cap={cap_range} key={score_context['selected_key']} seuils={seuil_achat_opt}/{seuil_vente_opt}")

                try:
                    sig, last_price, trend, last_rsi, volume_mean, score, derivatives = get_trading_signal(
                        prices, volumes, domaine=domaine, return_derivatives=True, symbol=symbol, cap_range=cap_range,
                        seuil_achat=seuil_achat_opt, seuil_vente=seuil_vente_opt,
                        price_extras=extras_to_use
                    )
                except Exception as e:
                    print(f"⚠️ Erreur get_trading_signal pour {symbol}: {e}")
                    import traceback
                    traceback.print_exc()
                    derivatives = {}

                # ✅ Backfill DB + dériver cap_range si manquant, puis recomputer si contexte change
                _ctx_changed = False
                try:
                    deriv_sector = derivatives.get('sector')
                    deriv_mc = float(derivatives.get('market_cap_val') or 0)
                    _need_update = False
                    _s = None
                    _c = None
                    _m = None
                    if deriv_sector and deriv_sector not in ('Inconnu', 'Unknown', '') and domaine in ('Inconnu', 'Unknown', '', None):
                        from sector_normalizer import normalize_sector
                        domaine = normalize_sector(deriv_sector)
                        _s = domaine
                        _need_update = True
                        _ctx_changed = True
                    if deriv_mc > 0:
                        _m = deriv_mc
                        if cap_range in ('Unknown', '', None):
                            cap_range = qsi.classify_cap_range(deriv_mc)
                            _ctx_changed = True
                        _c = cap_range
                        _need_update = True
                    if _need_update:
                        qsi.update_symbol_info_in_db(symbol, sector=_s, cap_range=_c, market_cap_b=_m)
                except Exception:
                    pass

                # ✅ Si le contexte a changé (secteur ou cap dérivé), recomputer signal+score
                if _ctx_changed:
                    score_context2 = qsi.resolve_symbol_scoring_context(
                        symbol, domaine=domaine, cap_range=cap_range, best_params=best_params_all,
                    )
                    domaine = score_context2['domaine']
                    cap_range = score_context2['cap_range']
                    seuil_achat_opt = score_context2['seuil_achat']
                    seuil_vente_opt = score_context2['seuil_vente']
                    extras_to_use = score_context2['price_extras']
                    self._debug_log(f"🔄 {symbol}: recompute ctx domaine={domaine} cap={cap_range} key={score_context2['selected_key']} seuils={seuil_achat_opt}/{seuil_vente_opt}")
                    try:
                        sig, last_price, trend, last_rsi, volume_mean, score, derivatives = get_trading_signal(
                            prices, volumes, domaine=domaine, return_derivatives=True, symbol=symbol, cap_range=cap_range,
                            seuil_achat=seuil_achat_opt, seuil_vente=seuil_vente_opt,
                            price_extras=extras_to_use
                        )
                    except Exception:
                        pass

                consensus_data = qsi.get_consensus(symbol) or {}

                row_info = {
                    'Symbole': symbol,
                    'Signal': sig,
                    'Score': score,
                    'Prix': last_price,
                    'Devise': str(stock_data.get('Currency', 'USD')),
                    'FxRateToUSD': float(stock_data.get('FxRateToUSD', 1.0) or 1.0),
                    'Tendance': 'Hausse' if trend else 'Baisse',
                    'RSI': last_rsi,
                    'DomaineOriginal': original_domaine,
                    'Domaine': domaine,
                    'CapRange': cap_range,
                    'Volume moyen': float(derivatives.get('volume_mean_usd', volume_mean * last_price)),
                    # Consensus (stable via cache/offline fallback)
                    'Consensus': consensus_data.get('label', 'Neutre'),
                    'ConsensusMean': consensus_data.get('mean', None),
                    'dPrice': round((derivatives.get('price_slope_rel') or 0.0) * 100, 2),
                    'Var5j (%)': round(float((derivatives.get('var_5j_pct') or 0.0)), 2),
                    'dRSI': round((derivatives.get('rsi_slope_rel') or 0.0) * 100, 2),
                    'dVolRel': round((derivatives['volume_slope_rel_usd'] if 'volume_slope_rel_usd' in derivatives else derivatives.get('volume_slope_rel', 0.0) or 0.0) * 100, 2),
                    # ✅ Métriques financières simples - protection contre None
                    'Rev. Growth (%)': round(float((derivatives.get('rev_growth_val') or 0.0)), 2),
                    'EBITDA Yield (%)': round(float((derivatives.get('ebitda_yield_pct') or 0.0)), 2),
                    'FCF Yield (%)': round(float((derivatives.get('fcf_yield_pct') or 0.0)), 2),
                    'EBITDA (B$)': round(float((derivatives.get('ebitda_val') or 0.0)), 2),
                    'FCF (B$)': round(float((derivatives.get('fcf_val') or 0.0)), 2),
                    'D/E Ratio': round(float((derivatives.get('debt_to_equity') or 0.0)), 2),
                    'Market Cap (B$)': round(float((derivatives.get('market_cap_val') or 0.0)), 2),
                    'ROE (%)': round(float((derivatives.get('roe_val') or 0.0)), 2)
                }

                self.current_results.append(row_info)
            except Exception as e:
                # ✅ Ne jamais ignorer silencieusement - ajouter au moins les données de base
                print(f"❌ Erreur critique pour {symbol}: {e}")
                try:
                    row_info = {
                        'Symbole': symbol,
                        'Signal': 'ERREUR',
                        'Score': 0.0,
                        'Prix': float(stock_data['Close'].iloc[-1]) if 'Close' in stock_data else 0.0,
                        'Devise': str(stock_data.get('Currency', 'USD')),
                        'FxRateToUSD': float(stock_data.get('FxRateToUSD', 1.0) or 1.0),
                        'Tendance': 'N/A',
                        'RSI': 0.0,
                        'Domaine': 'Inconnu',
                        'CapRange': cap_range if 'cap_range' in locals() and cap_range else 'Unknown',
                        'Volume moyen': 0.0,
                        'dPrice': 0.0,
                        'Var5j (%)': 0.0,
                        'dRSI': 0.0,
                        'dVolRel': 0.0,
                        'Rev. Growth (%)': 0.0,
                        'EBITDA Yield (%)': 0.0,
                        'FCF Yield (%)': 0.0,
                        'EBITDA (B$)': 0.0,
                        'FCF (B$)': 0.0,
                        'D/E Ratio': 0.0,
                        'Market Cap (B$)': 0.0,
                        'ROE (%)': 0.0
                    }
                    self.current_results.append(row_info)
                except Exception:
                    pass

        # Attach fiabilite AND nb_trades from backtests if present
        if backtests:
            bt_map = {b['Symbole']: b for b in backtests}
            for r in self.current_results:
                sym = r['Symbole']
                if sym in bt_map:
                    r['Fiabilite'] = bt_map[sym].get('taux_reussite', 'N/A')
                    r['NbTrades'] = bt_map[sym].get('trades', 0)
                else:
                    r['Fiabilite'] = 'N/A'
                    r['NbTrades'] = 0
        else:
            for r in self.current_results:
                r['Fiabilite'] = 'N/A'
                r['NbTrades'] = 0


        # Apply fiabilité and nb trades filters
        min_f = getattr(self, 'min_fiabilite_spin', None)
        min_t = getattr(self, 'min_trades_spin', None)
        include_non = getattr(self, 'include_none_val_chk', None)
        try:
            min_val = int(min_f.value()) if min_f is not None else 60
        except Exception:
            min_val = 60
        try:
            min_trades = int(min_t.value()) if min_t is not None else 5
        except Exception:
            min_trades = 5
        try:
            include_none_val = bool(include_non.isChecked()) if include_non is not None else True
        except Exception:
            include_none_val = True

        filtered = []
        for r in self.current_results:
            fiab = r.get('Fiabilite', 'N/A')
            nb_trades = r.get('NbTrades', 0)

            # Filtre sur nb trades - afficher TOUS les stocks inclus ceux avec 0 trades
            try:
                # Toujours inclure les résultats, même avec 0 trades
                # On affiche simplement l'indicateur 0 trades sans filtrer
                if int(nb_trades) > 0 and int(nb_trades) < min_trades:
                    # Si il y a des trades mais moins que le minimum, filtrer
                    if not include_none_val:
                        continue
            except Exception:
                # If we can't parse nb_trades, only include when include_none_val is True
                if not include_none_val:
                    continue

            # Filtre sur fiabilité
            try:
                if fiab == 'N/A':
                    if include_none_val:
                        filtered.append(r)
                else:
                    if float(fiab) >= float(min_val):
                        filtered.append(r)
            except Exception:
                if include_none_val:
                    filtered.append(r)

        self._schedule_result_visuals_refresh(result, mode='download')

        # NOTE: Do not replace the user's popular/mes lists with filtered results.
        # Instead, we can optionally update item tooltips to show fiabilité without
        # modifying the list contents. Keep original lists intact so the user
        # doesn't lose their configured popular symbols.
        try:
            # Map fiabilité by symbol for quick lookup
            fiab_map = {r['Symbole']: r.get('Fiabilite', 'N/A') for r in filtered}
            # Update tooltip for items in popular_list only (non-destructive)
            for i in range(self.popular_list.count()):
                item = self.popular_list.item(i)
                sym = item.data(Qt.UserRole) if item.data(Qt.UserRole) is not None else item.text()
                if sym in fiab_map:
                    item.setToolTip(f"Fiabilité: {fiab_map[sym]}")
                else:
                    item.setToolTip("")
            # Do the same for mes_list
            for i in range(self.mes_list.count()):
                item = self.mes_list.item(i)
                sym = item.data(Qt.UserRole) if item.data(Qt.UserRole) is not None else item.text()
                if sym in fiab_map:
                    item.setToolTip(f"Fiabilité: {fiab_map[sym]}")
                else:
                    item.setToolTip("")
            # And for coko_list
            if hasattr(self, 'coko_list'):
                for i in range(self.coko_list.count()):
                    item = self.coko_list.item(i)
                    sym = item.data(Qt.UserRole) if item.data(Qt.UserRole) is not None else item.text()
                    if sym in fiab_map:
                        item.setToolTip(f"Fiabilité: {fiab_map[sym]}")
                    else:
                        item.setToolTip("")
        except Exception:
            pass

        # Finalize results displayed in table
        self.update_results_table()
        # If backtest results present, render the backtest summary and table
        try:
            backtests = result.get('backtest_results', []) if isinstance(result, dict) else []
            if backtests:
                # current_results contains signal rows with Domaine if available
                self.render_backtest_summary_and_table(backtests, self.current_results)
        except Exception:
            pass

    def on_analysis_progress(self, message):
        if self.progress:
            self.progress.setLabelText(message)

    def on_analysis_complete(self, result):
        # 🔧 Vérifier que ce résultat appartient à l'analyse actuelle
        received_id = result.get('_analysis_id', 0) if isinstance(result, dict) else 0
        if received_id != self._analysis_id:
            print(f"⚠️ Résultat ignoré: ID={received_id}, ID actuel={self._analysis_id}")
            return

        self._analysis_running = False
        self._active_analysis_thread = None
        
        # Re-enable all buttons
        self.analyze_button.setEnabled(True)
        self.backtest_button.setEnabled(True)
        
        if self.progress:
            self.progress.close()
        
        # 🔧 Supprimer filtered_results si existant pour forcer l'utilisation de current_results
        if hasattr(self, 'filtered_results'):
            delattr(self, 'filtered_results')
        
        # Stocker les résultats
        self.current_results = result.get('signals', [])
        
        # 🔧 Stocker les résultats du backtest dans une map pour accès rapide
        backtest_results = result.get('backtest_results', []) if isinstance(result, dict) else []
        self.backtest_map = {b['Symbole']: b for b in backtest_results} if backtest_results else {}
        
        # 🔧 Ajouter les données de backtest aux signaux
        for signal in self.current_results:
            sym = signal.get('Symbole')
            if sym in self.backtest_map:
                bt = self.backtest_map[sym]
                signal['Fiabilite'] = bt.get('taux_reussite', 'N/A')
                signal['NbTrades'] = bt.get('trades', 0)
                signal['Gagnants'] = bt.get('gagnants', 0)
                signal['Gain_total'] = bt.get('gain_total', 0.0)
                signal['Gain_moyen'] = bt.get('gain_moyen', 0.0)
                signal['Drawdown_max'] = bt.get('drawdown_max', 0.0)
            else:
                signal.setdefault('Fiabilite', 'N/A')
                signal.setdefault('NbTrades', 0)
                signal.setdefault('Gagnants', 0)
                signal.setdefault('Gain_total', 0.0)
                signal.setdefault('Gain_moyen', 0.0)
                signal.setdefault('Drawdown_max', 0.0)
        
        # 🔧 Initialiser les colonnes par défaut pour tous les signaux
        for r in self.current_results:
            if not r.get('CapRange'):
                r['CapRange'] = 'Unknown'
            r.setdefault('Devise', 'USD')
            r.setdefault('FxRateToUSD', 1.0)
            r.setdefault('dPrice', 0.0)
            r.setdefault('Var5j (%)', 0.0)
            r.setdefault('dRSI', 0.0)
            r.setdefault('dVolRel', 0.0)
            r.setdefault('Rev. Growth (%)', 0.0)
            r.setdefault('EBITDA Yield (%)', 0.0)
            r.setdefault('FCF Yield (%)', 0.0)
            r.setdefault('EBITDA (B$)', 0.0)
            r.setdefault('FCF (B$)', 0.0)
            r.setdefault('D/E Ratio', 0.0)
            r.setdefault('Market Cap (B$)', 0.0)
            r.setdefault('ROE (%)', 0.0)
        
        # Charger les paramètres optimisés (lazy-load depuis SQLite si vide)
        self._get_best_parameters_cached()
        
        # ✅ OPTIMISATION: Réutiliser les données déjà téléchargées depuis result
        try:
            # Récupérer les données déjà disponibles
            existing_data = result.get('data', {}) if isinstance(result, dict) else {}
            
            for r in self.current_results:
                sym = r.get('Symbole')
                if not sym:
                    continue
                    
                # Calculer les dérivées techniques et métriques financières si manquantes
                need_derivatives = not r.get('dPrice') or float(r.get('dPrice', 0)) == 0.0
                need_financials = not r.get('Market Cap (B$)') or float(r.get('Market Cap (B$)', 0)) == 0.0
                if need_derivatives or need_financials:
                    try:
                        # Réutiliser strictement les données en mémoire pour éviter
                        # les appels réseau dans le callback UI final.
                        stock_data = existing_data.get(sym)
                        
                        if stock_data is None:
                            continue
                            
                        prices = stock_data['Close']
                        volumes = stock_data['Volume']
                        try:
                            score_context = qsi.resolve_symbol_scoring_context(
                                sym,
                                domaine=r.get('Domaine', 'Inconnu'),
                                cap_range=r.get('CapRange'),
                                best_params=self.best_parameters,
                            )
                            _sig, _last_price, _trend, _last_rsi, _vol_mean, _score, derivatives = get_trading_signal(
                                prices, volumes,
                                domaine=score_context['domaine'],
                                return_derivatives=True,
                                symbol=sym,
                                cap_range=score_context['cap_range'],
                                seuil_achat=score_context['seuil_achat'],
                                seuil_vente=score_context['seuil_vente'],
                                price_extras=score_context['price_extras'],
                            )
                        except Exception:
                            derivatives = {}

                        # Dérivées techniques (relatives en %)
                        if need_derivatives:
                            r['dPrice'] = round(derivatives.get('price_slope_rel', 0.0) * 100, 2)
                            r['Var5j (%)'] = round(float(derivatives.get('var_5j_pct', 0.0)), 2)
                            r['dRSI'] = round(derivatives.get('rsi_slope_rel', 0.0) * 100, 2)
                            r['dVolRel'] = round((derivatives['volume_slope_rel_usd'] if 'volume_slope_rel_usd' in derivatives else derivatives.get('volume_slope_rel', 0.0) or 0.0) * 100, 2)
                        
                        # ✅ Métriques financières simples
                        r['Rev. Growth (%)'] = round(derivatives.get('rev_growth_val', 0.0), 2)
                        r['EBITDA Yield (%)'] = round(derivatives.get('ebitda_yield_pct', 0.0), 2)
                        r['FCF Yield (%)'] = round(derivatives.get('fcf_yield_pct', 0.0), 2)
                        r['D/E Ratio'] = round(derivatives.get('debt_to_equity', 0.0), 2)
                        r['Market Cap (B$)'] = round(derivatives.get('market_cap_val', 0.0), 2)

                        # ✅ Dériver cap_range depuis market_cap si encore Unknown
                        mc_val = derivatives.get('market_cap_val')
                        if mc_val and float(mc_val) > 0 and r.get('CapRange') in ('Unknown', '', None):
                            r['CapRange'] = qsi.classify_cap_range(float(mc_val))
                    except Exception:
                        # leave defaults
                        if need_derivatives:
                            r.setdefault('dPrice', 0.0)
                            r.setdefault('Var5j (%)', 0.0)
                            r.setdefault('dRSI', 0.0)
                            r.setdefault('dVolRel', 0.0)
        except Exception:
            pass

        # Afficher les résultats
        self.update_results_table()
        self._schedule_result_visuals_refresh(result, mode='analysis')
        # Render backtest summary/table if present in the result
        try:
            backtests = result.get('backtest_results', []) if isinstance(result, dict) else []
            signals = result.get('signals', []) if isinstance(result, dict) else []
            if backtests:
                self.render_backtest_summary_and_table(backtests, signals)
        except Exception:
            pass

    def on_analysis_error(self, error_msg):
        self._analysis_running = False
        self._active_analysis_thread = None
        self.analyze_button.setEnabled(True)
        if self.progress:
            self.progress.close()
        
        QMessageBox.critical(self, "Erreur", f"Erreur pendant l'analyse:\n{error_msg}")

    def analyze_stock(self):
        if self._analysis_running:
            QMessageBox.information(self, "Analyse en cours", "Une analyse est déjà en cours. Attends la fin avant d'en lancer une autre.")
            return

        # Get list of symbols from input or from selection in lists
        symbols = [s.strip().upper() for s in self.symbol_input.text().split(",") if s.strip()]
        if not symbols:
            # If no manual input, use selected items from the lists (popular first, then mes)
            selected = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in self.popular_list.selectedItems()]
            if not selected:
                selected = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in self.mes_list.selectedItems()]
            if not selected and hasattr(self, 'coko_list'):
                selected = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in self.coko_list.selectedItems()]
            symbols = [s.strip().upper() for s in selected if s]
        if not symbols:
            QMessageBox.warning(self, "Erreur", "Veuillez entrer au moins un symbole")
            return

        # Get analysis period
        period = self.period_input.currentData()
        if not period:
            QMessageBox.warning(self, "Erreur", "Veuillez sélectionner une période d'analyse")
            return

        # 🔧 Incrémenter l'ID d'analyse et stopper les threads précédents
        self._analysis_id += 1
        current_id = self._analysis_id
        print(f"\n🚀 Nouvelle analyse lancée #ID={current_id}")
        
        # 🔧 Réinitialiser les données complètement
        self.current_results = []
        if hasattr(self, 'filtered_results'):
            delattr(self, 'filtered_results')
        if hasattr(self, 'backtest_map'):
            delattr(self, 'backtest_map')
        
        # Disable buttons during analysis
        self.analyze_button.setEnabled(False)
        self.backtest_button.setEnabled(False)

        # Progress dialog
        self.progress = QProgressDialog("Téléchargement et analyse...", "Annuler", 0, 0, self)
        self.progress.setWindowTitle("Analyse")
        self.progress.setWindowModality(Qt.WindowModal)
        self.progress.setMinimumDuration(0)
        self.progress.setAutoClose(False)
        self.progress.setMinimumWidth(400)

        # Launch download thread (no backtest)
        min_holding_days = self.min_hold_days_spin.value() if hasattr(self, 'min_hold_days_spin') else 7
        self.download_thread = DownloadThread(
            symbols,
            period,
            analysis_id=current_id,
        )
        self.download_thread.result_ready.connect(self.on_download_complete)
        self.download_thread.error.connect(self.on_analysis_error)
        self.download_thread.progress.connect(self.on_analysis_progress)
        self._analysis_running = True
        self._active_analysis_thread = self.download_thread
        self.download_thread.start()
        print(f"📥 Download thread démarré avec ID={current_id}")

    def analyse_and_backtest(self):
        if self._analysis_running:
            QMessageBox.information(self, "Analyse en cours", "Une analyse est déjà en cours. Attends la fin avant d'en lancer une autre.")
            return

        # For consistency with 'Analyser mouvements fiables', run the full
        # analyse_signaux_populaires pipeline (which includes backtests) and
        # embed the same charts + detailed backtest info in the UI.

        # Get list of symbols from input or from selection in lists
        symbols = [s.strip().upper() for s in self.symbol_input.text().split(",") if s.strip()]
        if not symbols:
            selected = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in self.popular_list.selectedItems()]
            if not selected:
                selected = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in self.mes_list.selectedItems()]
            if not selected and hasattr(self, 'coko_list'):
                selected = [it.data(Qt.UserRole) if it.data(Qt.UserRole) is not None else it.text() for it in self.coko_list.selectedItems()]
            symbols = [s.strip().upper() for s in selected if s]
        if not symbols:
            QMessageBox.warning(self, "Erreur", "Veuillez entrer au moins un symbole")
            return

        period = self.period_input.currentData()
        if not period:
            QMessageBox.warning(self, "Erreur", "Veuillez sélectionner une période d'analyse")
            return

        # 🔧 Incrémenter l'ID d'analyse et stopper les threads précédents
        self._analysis_id += 1
        current_id = self._analysis_id
        print(f"\n🚀 Nouvelle analyse backtest lancée #ID={current_id}")
        
        # 🔧 Réinitialiser les données complètement
        self.current_results = []
        if hasattr(self, 'filtered_results'):
            delattr(self, 'filtered_results')
        if hasattr(self, 'backtest_map'):
            delattr(self, 'backtest_map')
        
        # Disable buttons during analysis
        self.analyze_button.setEnabled(False)
        self.backtest_button.setEnabled(False)

        # Progress dialog
        self.progress = QProgressDialog("Analyse et backtest en cours...", "Annuler", 0, 0, self)
        self.progress.setWindowTitle("Analyse + Backtest")
        self.progress.setWindowModality(Qt.WindowModal)
        self.progress.setMinimumDuration(0)
        self.progress.setAutoClose(False)
        self.progress.setMinimumWidth(400)

        # Use the AnalysisThread which calls analyse_signaux_populaires (no plt.show()
        # in background). Pass the selected symbols as the "popular_symbols" input
        # so the function analyzes/backtests those symbols and returns the same
        # result structure used for the "Analyser mouvements fiables" flow.
        selected_pop = symbols
        selected_mes = []

        min_holding_days = self.min_hold_days_spin.value() if hasattr(self, 'min_hold_days_spin') else 7
        self.analysis_thread = AnalysisThread(selected_pop, selected_mes, period, analysis_id=current_id, min_holding_days=min_holding_days)
        self.analysis_thread.result_ready.connect(self.on_analysis_complete)
        self.analysis_thread.error.connect(self.on_analysis_error)
        self.analysis_thread.progress.connect(self.on_analysis_progress)
        self._analysis_running = True
        self._active_analysis_thread = self.analysis_thread
        self.analysis_thread.start()
        print(f"📊 Analysis backtest thread démarré avec ID={current_id}")

    def update_results_table(self):
        """Fill the merged table (`self.merged_table`) with current results plus backtest metrics."""
        if not hasattr(self, 'current_results') or not hasattr(self, 'merged_table'):
            return

        if getattr(self, '_updating_results_table', False):
            return

        def _parse_numeric(val, default=None):
            try:
                if val is None:
                    return default
                if isinstance(val, str):
                    raw = val.strip()
                    if raw == '' or raw.upper() == 'N/A':
                        return default
                    raw = raw.replace('%', '').replace('$', '').replace(',', '').replace('x', '').strip()
                    if raw == '':
                        return default
                    return float(raw)
                return float(val)
            except Exception:
                return default

        def _set_item(row: int, col: int, value, *, numeric: bool = False):
            item = QTableWidgetItem(str(value))
            if numeric:
                try:
                    item.setData(2, float(value))
                except Exception:
                    pass
            self.merged_table.setItem(row, col, item)

        def _colorize(item, kind: str, value):
            try:
                if kind == 'signal':
                    text = str(value).lower()
                    if 'buy' in text or 'achat' in text:
                        item.setForeground(QColor(34, 139, 34))
                    elif 'sell' in text or 'vente' in text:
                        item.setForeground(QColor(255, 0, 0))
                    else:
                        item.setForeground(QColor(255, 165, 0))
                elif kind == 'fiab':
                    v = float(value)
                    if v >= 75:
                        item.setForeground(QColor(0, 128, 0))
                    elif v >= 50:
                        item.setForeground(QColor(34, 139, 34))
                    elif v >= 30:
                        item.setForeground(QColor(255, 165, 0))
                    else:
                        item.setForeground(QColor(255, 0, 0))
                elif kind in {'good_high', 'gain', 'positive'}:
                    v = float(value)
                    if v > 0:
                        item.setForeground(QColor(34, 139, 34))
                    elif v < 0:
                        item.setForeground(QColor(255, 0, 0))
                    else:
                        item.setForeground(QColor(255, 165, 0))
                elif kind == 'ratio_low':
                    v = float(value)
                    if v < 0.5:
                        item.setForeground(QColor(0, 128, 0))
                    elif v < 1.5:
                        item.setForeground(QColor(34, 139, 34))
                    elif v < 2.5:
                        item.setForeground(QColor(255, 165, 0))
                    else:
                        item.setForeground(QColor(255, 0, 0))
                elif kind == 'trend':
                    if str(value).lower().startswith('hausse'):
                        item.setForeground(QColor(34, 139, 34))
                    elif str(value).lower().startswith('baisse'):
                        item.setForeground(QColor(255, 0, 0))
                    else:
                        item.setForeground(QColor(255, 165, 0))
            except Exception:
                pass

        self._updating_results_table = True
        sorting_was_enabled = self.merged_table.isSortingEnabled()
        self.merged_table.setUpdatesEnabled(False)
        try:
            if sorting_was_enabled:
                self.merged_table.setSortingEnabled(False)

            min_fiab_threshold = self.fiab_threshold_spin.value() if hasattr(self, 'fiab_threshold_spin') else 30
            bt_map = getattr(self, 'backtest_map', {}) or {}
            results_to_display = []

            for result in self.current_results:
                if not isinstance(result, dict):
                    continue
                fiab_val = result.get('Fiabilite', 'N/A')
                if fiab_val == 'N/A':
                    results_to_display.append(result)
                    continue
                fiab_num = _parse_numeric(fiab_val, None)
                if fiab_num is None or fiab_num >= min_fiab_threshold:
                    results_to_display.append(result)

            self.merged_table.setRowCount(len(results_to_display))

            for row, signal in enumerate(results_to_display):
                sym = str(signal.get('Symbole', '')).strip()
                if not sym:
                    continue

                bt = bt_map.get(sym, {})
                score = _parse_numeric(signal.get('Score', 0.0), 0.0) or 0.0
                seuil_achat = _parse_numeric(signal.get('Seuil_Achat', signal.get('seuil_achat', 4.2)), 4.2) or 4.2
                seuil_vente = _parse_numeric(signal.get('Seuil_Vente', signal.get('seuil_vente', -0.5)), -0.5) or -0.5

                fiab = signal.get('Fiabilite', bt.get('taux_reussite', 'N/A') if bt else 'N/A')
                nb_trades = signal.get('NbTrades', bt.get('trades', 0) if bt else 0)
                gagnants = signal.get('Gagnants', bt.get('gagnants', 0) if bt else 0)
                gain_total = signal.get('Gain_total', bt.get('gain_total', 0.0) if bt else 0.0)
                gain_moyen = signal.get('Gain_moyen', bt.get('gain_moyen', 0.0) if bt else 0.0)

                values = {
                    0: sym,
                    1: signal.get('Signal', 'N/A'),
                    2: signal.get('Score', 0.0),
                    3: signal.get('Prix', 0.0),
                    4: signal.get('Tendance', 'N/A'),
                    5: signal.get('RSI', 0.0),
                    6: signal.get('Volume moyen', 0.0),
                    7: signal.get('Domaine', 'Inconnu'),
                    8: signal.get('CapRange', 'Unknown'),
                    9: score / seuil_achat if score >= 0 and seuil_achat else (score / seuil_vente if score < 0 and seuil_vente else 0.0),
                    10: fiab,
                    11: nb_trades,
                    12: gagnants,
                    13: signal.get('Rev. Growth (%)', 0.0),
                    14: signal.get('EBITDA Yield (%)', 0.0),
                    15: signal.get('FCF Yield (%)', 0.0),
                    16: signal.get('D/E Ratio', 0.0),
                    17: signal.get('Market Cap (B$)', 0.0),
                    18: signal.get('ROE (%)', 0.0),
                    19: signal.get('dPrice', 0.0),
                    20: signal.get('Var5j (%)', 0.0),
                    21: signal.get('dRSI', 0.0),
                    22: signal.get('dVolRel', 0.0),
                    23: gain_total,
                    24: gain_moyen,
                    25: signal.get('Consensus', 'N/A'),
                }

                numeric_cols = {2, 3, 5, 6, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24}
                for col, value in values.items():
                    _set_item(row, col, value, numeric=col in numeric_cols)

                _colorize(self.merged_table.item(row, 1), 'signal', values[1])
                _colorize(self.merged_table.item(row, 4), 'trend', values[4])
                _colorize(self.merged_table.item(row, 10), 'fiab', values[10] if values[10] != 'N/A' else 0)
                _colorize(self.merged_table.item(row, 13), 'gain', values[13])
                _colorize(self.merged_table.item(row, 14), 'gain', values[14])
                _colorize(self.merged_table.item(row, 15), 'gain', values[15])
                _colorize(self.merged_table.item(row, 16), 'ratio_low', values[16])
                _colorize(self.merged_table.item(row, 18), 'gain', values[18])
                _colorize(self.merged_table.item(row, 19), 'positive', values[19])
                _colorize(self.merged_table.item(row, 20), 'positive', values[20])
                _colorize(self.merged_table.item(row, 21), 'positive', values[21])
                _colorize(self.merged_table.item(row, 22), 'positive', values[22])
                _colorize(self.merged_table.item(row, 23), 'gain', values[23])
                _colorize(self.merged_table.item(row, 24), 'gain', values[24])

                consensus_item = self.merged_table.item(row, 25)
                if consensus_item is not None:
                    consensus_lower = str(values[25]).lower()
                    if 'strong buy' in consensus_lower or 'achat fort' in consensus_lower:
                        consensus_item.setForeground(QColor(0, 128, 0))
                    elif 'buy' in consensus_lower or 'achat' in consensus_lower:
                        consensus_item.setForeground(QColor(34, 139, 34))
                    elif 'hold' in consensus_lower or 'conserver' in consensus_lower or 'neutre' in consensus_lower:
                        consensus_item.setForeground(QColor(255, 165, 0))
                    elif 'sell' in consensus_lower or 'vente' in consensus_lower:
                        consensus_item.setForeground(QColor(255, 0, 0))

                if row == 0:
                    pass

            self._charts_dirty = True
            self._comparisons_dirty = True

            if hasattr(self, 'charts_container'):
                self._schedule_charts_refresh()
            if hasattr(self, 'comparisons_container'):
                self._schedule_comparisons_refresh()
        finally:
            if sorting_was_enabled:
                self.merged_table.setSortingEnabled(True)
            self.merged_table.setUpdatesEnabled(True)
            self._updating_results_table = False
    
    def toggle_bottom(self, checked: bool):
        """Hide/show the bottom summary/backtest/results panel."""
        if checked:
            # hide bottom container and expand plots
            if hasattr(self, 'bottom_container'):
                self.bottom_container.setVisible(False)
            self.toggle_bottom_btn.setText("Afficher détails")
            try:
                self.splitter.setSizes([self.height(), 0])
            except Exception:
                pass
        else:
            if hasattr(self, 'bottom_container'):
                self.bottom_container.setVisible(True)
            self.toggle_bottom_btn.setText("Masquer détails")
            try:
                total_h = max(600, self.height())
                top_h = int(total_h * 0.72)
                bottom_h = total_h - top_h
                self.splitter.setSizes([top_h, bottom_h])
            except Exception:
                pass
    
    def toggle_offline_mode(self):
        """Bascule entre le mode online et offline"""
        is_offline = self.offline_mode_btn.isChecked()
        qsi.OFFLINE_MODE = is_offline
        
        if is_offline:
            self.offline_mode_btn.setText("📴 Mode: OFFLINE")
            self.offline_mode_btn.setStyleSheet("QPushButton { background-color: #FF9800; color: white; font-weight: bold; }")
            self.summary_text.append("\n⚠️ Mode OFFLINE activé - Utilisation du cache uniquement")
        else:
            self.offline_mode_btn.setText("🌐 Mode: ONLINE")
            self.offline_mode_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
            self.summary_text.append("\n✅ Mode ONLINE activé - Téléchargement si cache obsolète")

    def toggle_debug_mode(self):
        """Active/désactive les logs debug en boucle."""
        self.debug_mode_enabled = self.debug_mode_btn.isChecked()
        if self.debug_mode_enabled:
            self.debug_mode_btn.setText("🐞 Debug: ON")
            self.debug_mode_btn.setStyleSheet("QPushButton { background-color: #E65100; color: white; font-weight: bold; }")
            self.summary_text.append("\n🐞 Mode DEBUG activé - logs détaillés affichés")
        else:
            self.debug_mode_btn.setText("🐞 Debug: OFF")
            self.debug_mode_btn.setStyleSheet("QPushButton { background-color: #9E9E9E; color: white; font-weight: bold; }")
            self.summary_text.append("\n✅ Mode DEBUG désactivé - logs en boucle masqués")

    def render_backtest_summary_and_table(self, backtest_results: list, signals: list):
        """Build the summary text and populate an internal backtest map used by the merged table.

        This no longer writes to a separate backtest table; instead it stores results in
        `self.backtest_map` and injects Fiabilité/NbTrades back into `self.current_results` so
        the merged table can display and sort them.
        """
        try:
            total_trades = 0
            total_gagnants = 0
            total_gain = 0.0

            # Map symbol to domain from signals list
            domain_map = {s.get('Symbole'): s.get('Domaine', 'Inconnu') for s in signals if isinstance(s, dict)}

            domain_stats = {}

            # Build backtest map
            self.backtest_map = {}
            for br in backtest_results:
                sym = br.get('Symbole')
                trades = int(br.get('trades', 0))
                gagnants = int(br.get('gagnants', 0))
                gain_total = float(br.get('gain_total', 0.0))
                gain_moyen = float(br.get('gain_moyen', 0.0))
                #drawdown = float(br.get('drawdown_max', 0.0))
                taux = float(br.get('taux_reussite', 0.0))

                self.backtest_map[sym] = {
                    'trades': trades,
                    'gagnants': gagnants,
                    'taux_reussite': taux,
                    'gain_total': gain_total,
                    'gain_moyen': gain_moyen,
                    #'drawdown_max': drawdown
                }

                total_trades += trades
                total_gagnants += gagnants
                total_gain += gain_total

                # Domain aggregation
                domain = domain_map.get(sym, 'Inconnu')
                if domain not in domain_stats:
                    domain_stats[domain] = {'trades': 0, 'gagnants': 0, 'gain': 0.0}
                domain_stats[domain]['trades'] += trades
                domain_stats[domain]['gagnants'] += gagnants
                domain_stats[domain]['gain'] += gain_total

            taux_global = (total_gagnants / total_trades * 100) if total_trades else 0.0

            # Inject Fiabilite/NbTrades into current_results so merged table shows them
            try:
                for r in getattr(self, 'current_results', []):
                    sym = r.get('Symbole')
                    bt = self.backtest_map.get(sym)
                    if bt:
                        r['Fiabilite'] = bt.get('taux_reussite', 'N/A')
                        r['NbTrades'] = bt.get('trades', 0)
                    else:
                        r.setdefault('Fiabilite', 'N/A')
                        r.setdefault('NbTrades', 0)
            except Exception:
                pass

            # Build summary text
            lines = []
            min_hold_days = self.min_hold_days_spin.value() if hasattr(self, 'min_hold_days_spin') else 7
            lines.append(
                f"🌍 Résultat global :\n"
                f" - Taux de réussite = {taux_global:.1f}%\n"
                f" - Nombre de trades = {total_trades}\n"
                f" - Gain total brut = {total_gain:.2f} $\n"
                f" - Durée min position = {min_hold_days} jour(s) actif(s)"
            )
            lines.append("\n📊 Taux de réussite par domaine:")
            for dom, stats in sorted(domain_stats.items(), key=lambda x: -x[1]['trades']):
                trades = stats['trades']
                gagnants = stats['gagnants']
                taux = (gagnants / trades * 100) if trades > 0 else 0.0
                gain_dom = stats['gain']
                lines.append(f" - {dom}: Trades={trades} | Gagnants={gagnants} | Taux={taux:.1f}% | Gain brut={gain_dom:.2f} $")

            self.summary_text.setPlainText('\n'.join(lines))

        except Exception:
            try:
                self.summary_text.setPlainText('')
            except Exception:
                pass
    
    def closeEvent(self, event):
        # Restore standard streams before widgets are torn down.
        try:
            sys.stdout = sys.__stdout__
            sys.stderr = sys.__stderr__
        except Exception:
            pass

        # Stop analysis thread if running
        try:
            if hasattr(self, 'analysis_thread') and self.analysis_thread.isRunning():
                self.analysis_thread.stop()
                self.analysis_thread.wait(2000)
        except (RuntimeError, AttributeError):
            pass
        try:
            if hasattr(self, 'download_thread') and self.download_thread.isRunning():
                self.download_thread.stop()
                self.download_thread.wait(2000)
        except (RuntimeError, AttributeError):
            pass

        try:
            if _CRASH_LOG_FILE is not None:
                _CRASH_LOG_FILE.flush()
        except Exception:
            pass
        event.accept()

    def clear_plots(self):
        for i in reversed(range(self.plots_layout.count())):
            w = self.plots_layout.itemAt(i).widget()
            if w:
                if hasattr(w, 'figure'):
                    w.figure.clear()
                    try:
                        w.close()
                    except Exception:
                        pass
                w.setParent(None)
        import gc
        gc.collect()

    def _compute_score_series(self, prices, volumes, domaine='Inconnu', cap_range=None, symbol=None):
        """Calcule l'evolution du score sur la fenetre analysee (jours actifs).
        
        ✅ THREAD-SAFE: Garantit que cap_range reste constant tout au long de la calcul
        """
        score_dates = []
        score_values = []
        start_idx = 50

        if len(prices) <= start_idx:
            return score_dates, score_values

        # Résoudre une seule fois le contexte exact du symbole
        score_context = qsi.resolve_symbol_scoring_context(
            symbol or '',
            domaine=domaine,
            cap_range=cap_range,
            best_params=self._get_best_parameters_cached(),
        )
        original_cap_range = score_context['cap_range']
        original_domaine = score_context['domaine']
        seuil_achat = score_context['seuil_achat']
        seuil_vente = score_context['seuil_vente']
        price_extras = score_context['price_extras']

        # Precision maximale: calcul quotidien (chaque jour actif) pour eviter
        # les artefacts visuels d'interpolation sur les seuils.
        step = 1

        for i in range(start_idx, len(prices), step):
            try:
                _sig, _last_price, _trend, _last_rsi, _vol_mean, score, derivatives = get_trading_signal(
                    prices.iloc[:i + 1],
                    volumes.iloc[:i + 1],
                    domaine=original_domaine,  # ✅ Toujours utiliser ORIGINAL
                    cap_range=original_cap_range,  # ✅ Toujours utiliser ORIGINAL
                    symbol=symbol,
                    seuil_achat=seuil_achat,
                    seuil_vente=seuil_vente,
                    price_extras=price_extras,
                    return_derivatives=True,
                )
                # 🔍 Assertion de sécurité
                used_cap = derivatives.get('_cap_range_used', original_cap_range)
                if used_cap != original_cap_range and original_cap_range is not None:
                    self._debug_log(f"⚠️ {symbol}: cap_range décalé de {original_cap_range} à {used_cap} à l'itération {i}")
                
                score_dates.append(prices.index[i])
                score_values.append(float(score))
            except Exception:
                continue

        # Assurer un point final (dernier jour) pour la lecture visuelle
        if score_dates and score_dates[-1] != prices.index[-1]:
            try:
                _sig, _last_price, _trend, _last_rsi, _vol_mean, score, derivatives = get_trading_signal(
                    prices,
                    volumes,
                    domaine=original_domaine,
                    cap_range=original_cap_range,
                    symbol=symbol,
                    seuil_achat=seuil_achat,
                    seuil_vente=seuil_vente,
                    price_extras=price_extras,
                    return_derivatives=True,
                )
                score_dates.append(prices.index[-1])
                score_values.append(float(score))
            except Exception:
                pass

        return score_dates, score_values

    def _get_global_thresholds_for_symbol(self, domaine='Inconnu', cap_range=None):
        """Retourne (seuil_achat, seuil_vente) optimises pour secteur/cap, avec fallback par defaut.
        
        ✅ GARANTIE: Les seuils retournés correspondent EXACTEMENT au cap_range fourni
        """
        default_buy = 4.2
        default_sell = -0.5
        try:
            best_params = self._get_best_parameters_cached()
            selected_key = None

            # ✅ ORDRE CRITIQUE: Chercher d'abord cap_range+domaine, sinon juste domaine
            if cap_range and cap_range != 'Unknown':
                comp_key = f"{domaine}_{cap_range}"
                if comp_key in best_params:
                    selected_key = comp_key

            if not selected_key and domaine in best_params:
                selected_key = domaine

            if selected_key:
                _coeffs, _thresholds, globals_thresholds, _gain, _extras = best_params[selected_key]
                buy_thr = float(globals_thresholds[0])
                sell_thr = float(globals_thresholds[1])
                
                # 🔍 Debug log pour tracer les seuils utilisés
                if domaine in ['Real Estate', 'Utilities'] or (cap_range and cap_range != 'Unknown'):
                    self._debug_log(f"✅ SEUILS: cap_range={cap_range}, key={selected_key} → buy={buy_thr:.2f}, sell={sell_thr:.2f}")
                
                return buy_thr, sell_thr
        except Exception as e:
            print(f"⚠️ Erreur _get_global_thresholds_for_symbol: {e}")

        return default_buy, default_sell

    def _build_symbol_figure_with_score(self, sym, prices, volumes, precomp=None, events=None):
        """Construit une figure: trace principal + score au fil du temps en dessous."""
        precomp = precomp or {}
        events = events or []

        fig = Figure(figsize=(10, 8.4))
        gs = fig.add_gridspec(2, 1, height_ratios=[3.2, 1.25], hspace=0.20)
        ax_main = fig.add_subplot(gs[0, 0])
        ax_score = fig.add_subplot(gs[1, 0], sharex=ax_main)

        score_dates = precomp.get('score_dates') or []
        score_values = precomp.get('score_values') or []
        # Toujours recalculer: les scores du backtest utilisent des fondamentaux
        # point-in-time (PIT) qui diffèrent du pickle cache utilisé par le score
        # affiché dans le titre.  Recalculer garantit la cohérence visuelle.
        score_dates, score_values = self._compute_score_series(
            prices,
            volumes,
            domaine=precomp.get('domaine', 'Inconnu'),
            cap_range=precomp.get('cap_range'),
            symbol=sym,
        )

        score_val = precomp.get('score')

        try:
            # Use show_xaxis=True to avoid plot_unified_chart clearing shared x tick labels.
            plot_unified_chart(sym, prices, volumes, ax_main, show_xaxis=True, score_override=score_val, precomputed=precomp)
        except Exception:
            ax_main.plot(prices.index, prices.values, color='black', linewidth=1.2)
            if isinstance(score_val, (int, float)):
                ax_main.set_title(f"{sym} | Score: {score_val:.2f}")
            else:
                ax_main.set_title(sym)

        # Hide x labels on top panel only; keep bottom panel date labels visible.
        ax_main.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

        buy_thr = precomp.get('seuil_achat')
        sell_thr = precomp.get('seuil_vente')
        if buy_thr is None or sell_thr is None:
            buy_thr, sell_thr = self._get_global_thresholds_for_symbol(
                domaine=precomp.get('domaine', 'Inconnu'),
                cap_range=precomp.get('cap_range'),
            )
        else:
            try:
                buy_thr = float(buy_thr)
                sell_thr = float(sell_thr)
            except Exception:
                buy_thr, sell_thr = self._get_global_thresholds_for_symbol(
                    domaine=precomp.get('domaine', 'Inconnu'),
                    cap_range=precomp.get('cap_range'),
                )

        for ev in events:
            if ev.get('type') == 'BUY':
                ax_main.scatter(ev['date'], ev['price'], marker='^', s=80, color='green', edgecolor='black', zorder=6)
            elif ev.get('type') == 'SELL':
                ax_main.scatter(ev['date'], ev['price'], marker='v', s=80, color='red', edgecolor='black', zorder=6)

        if score_dates and score_values:
            ax_score.plot(score_dates, score_values, color='#1565C0', linewidth=1.6, label='Score')
            ax_score.axhline(y=buy_thr, color='green', linestyle='--', alpha=0.5, linewidth=1.0, label=f'Seuil Achat ({buy_thr:.2f})')
            ax_score.axhline(y=sell_thr, color='red', linestyle='--', alpha=0.5, linewidth=1.0, label=f'Seuil Vente ({sell_thr:.2f})')

            # Affiche les BUY/SELL directement sur la courbe de score pour garantir
            # la correspondance visuelle entre positions et historique de score.
            try:
                idx = pd.to_datetime(score_dates, errors='coerce')
                score_index = pd.Index(idx)
                score_vals = list(score_values)
                for ev in events:
                    ev_type = str(ev.get('type', '')).upper()
                    ev_date = ev.get('date')
                    if not ev_date or ev_type not in {'BUY', 'SELL'}:
                        continue
                    ts = pd.to_datetime(ev_date, errors='coerce')
                    if pd.isna(ts) or score_index.empty:
                        continue
                    pos = score_index.get_indexer([ts], method='nearest')[0]
                    if pos < 0 or pos >= len(score_vals):
                        continue
                    y = float(score_vals[pos])
                    marker = '^' if ev_type == 'BUY' else 'v'
                    color = 'green' if ev_type == 'BUY' else 'red'
                    ax_score.scatter(score_dates[pos], y, marker=marker, s=48, color=color, edgecolor='black', zorder=7)
            except Exception:
                pass

            ax_score.legend(loc='upper left', fontsize=8, frameon=True)
        else:
            ax_score.text(0.5, 0.5, 'Score indisponible', transform=ax_score.transAxes,
                          ha='center', va='center', fontsize=9)

        ax_score.set_ylabel('Score', fontsize=9)
        ax_score.set_xlabel('Date', fontsize=9)
        ax_score.text(
            0.01,
            0.98,
            'Seuils appliques au score (pas au prix)',
            transform=ax_score.transAxes,
            va='top',
            ha='left',
            fontsize=8,
            color='#424242',
            bbox=dict(facecolor='white', alpha=0.6, edgecolor='none', pad=1.5),
        )
        ax_score.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=4, maxticks=8))
        ax_score.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax_score.xaxis.get_major_locator()))
        ax_score.tick_params(axis='x', labelrotation=0, labelsize=8)
        ax_score.grid(True, alpha=0.25)
        # Evite le warning Matplotlib avec axes jumeles (plot_unified_chart utilise twinx)
        # et réserve plus d'espace en haut pour les titres longs.
        fig.subplots_adjust(left=0.07, right=0.93, top=0.90, bottom=0.08, hspace=0.24)
        return fig




    def compute_domain_stats(self):
        """Agrège les résultats par domaine depuis la merged_table."""
        try:
            if not hasattr(self, 'merged_table') or self.merged_table.rowCount() == 0:
                return {'global': {}, 'by_domain': {}}
            
            domain_stats = {}
            total_trades = 0
            total_gagnants = 0
            total_gain = 0.0
            
            for row in range(self.merged_table.rowCount()):
                try:
                    # Colonne 7: Domaine
                    domaine_item = self.merged_table.item(row, 7)
                    domaine = domaine_item.text() if domaine_item and domaine_item.text().strip() else 'Inconnu'
                    
                    # Colonne 11: Nb Trades
                    trades_item = self.merged_table.item(row, 11)
                    nb_trades = int(trades_item.data(Qt.EditRole)) if trades_item and trades_item.data(Qt.EditRole) is not None else 0
                    
                    # Colonne 12: Gagnants
                    gagnants_item = self.merged_table.item(row, 12)
                    gagnants = int(gagnants_item.data(Qt.EditRole)) if gagnants_item and gagnants_item.data(Qt.EditRole) is not None else 0
                    
                    # Colonne 23: Gain total ($)
                    gain_item = self.merged_table.item(row, 23)
                    gain = float(gain_item.data(Qt.EditRole)) if gain_item and gain_item.data(Qt.EditRole) is not None else 0.0
                    
                    if domaine not in domain_stats:
                        domain_stats[domaine] = {'trades': 0, 'gagnants': 0, 'gain': 0.0}
                    
                    domain_stats[domaine]['trades'] += nb_trades
                    domain_stats[domaine]['gagnants'] += gagnants
                    domain_stats[domaine]['gain'] += gain
                    
                    total_trades += nb_trades
                    total_gagnants += gagnants
                    total_gain += gain
                except Exception as e:
                    print(f"⚠️ Erreur compute_domain_stats row {row}: {e}")
                    continue
            
            # Calculer taux de réussite par domaine
            for domaine in domain_stats:
                trades = domain_stats[domaine]['trades']
                gagnants = domain_stats[domaine]['gagnants']
                taux = (gagnants / trades * 100) if trades > 0 else 0.0
                domain_stats[domaine]['taux'] = taux
            
            total_taux = (total_gagnants / total_trades * 100) if total_trades > 0 else 0.0
            
            return {
                'global': {
                    'trades': total_trades,
                    'gagnants': total_gagnants,
                    'taux': total_taux,
                    'gain': total_gain
                },
                'by_domain': domain_stats
            }
        except Exception as e:
            print(f"❌ Erreur compute_domain_stats: {e}")
            import traceback
            traceback.print_exc()
            return {'global': {}, 'by_domain': {}}
    
    def populate_charts_tab(self):
        """Génère les graphiques de comparaison par domaine dans l'onglet Graphiques."""
        try:
            # Nettoyer l'onglet Graphiques
            while self.charts_scroll_layout.count() > 0:
                widget = self.charts_scroll_layout.takeAt(0).widget()
                if widget:
                    widget.deleteLater()
            
            stats = self.compute_domain_stats()
            if not stats['by_domain']:
                self.charts_scroll_layout.addWidget(QLabel("Aucun résultat à afficher. Analysez d'abord des symboles."))
                return
            
            # Titre + résumé global compact (sur une seule ligne)
            title = QLabel("📊 Analyse par domaine")
            title.setStyleSheet("font-weight: bold; font-size: 12px;")
            
            global_info = stats['global']
            min_hold_days = self.min_hold_days_spin.value() if hasattr(self, 'min_hold_days_spin') else 7
            summary_text = (
                f"🌍 Résultat global: Taux={global_info['taux']:.1f}% | Trades={global_info['trades']} | "
                f"Gain=${global_info['gain']:.2f} | Durée min={min_hold_days}j"
            )
            summary_label = QLabel(summary_text)
            summary_label.setStyleSheet("background-color: #f0f0f0; padding: 6px; border-radius: 4px; font-size: 10px;")
            
            # Graphiques matplotlib (sans tableau textuel)
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            
            # Préparer les données pour les graphiques
            domains = list(stats['by_domain'].keys())
            taux_list = [stats['by_domain'][d]['taux'] for d in domains]
            trades_list = [stats['by_domain'][d]['trades'] for d in domains]
            gain_list = [stats['by_domain'][d]['gain'] for d in domains]
            
            # Compter le nombre de symboles par domaine
            symbols_per_domain = {}
            for row in range(self.merged_table.rowCount()):
                try:
                    domaine_item = self.merged_table.item(row, 7)
                    domaine = domaine_item.text() if domaine_item else 'Inconnu'
                    symbols_per_domain[domaine] = symbols_per_domain.get(domaine, 0) + 1
                except Exception:
                    continue
            
            symbols_list = [symbols_per_domain.get(d, 0) for d in domains]
            
            # Calculer la rentabilité annualisée en % par secteur
            # Capital investi: 50€ par stock
            period_str = self.period_input.currentData() if hasattr(self, 'period_input') else "15mo"
            
            # Convertir la période en années
            if 'y' in period_str:
                years = float(period_str.replace('y', ''))
            elif 'mo' in period_str:
                months = float(period_str.replace('mo', ''))
                years = months / 12.0
            elif 'd' in period_str:
                days = float(period_str.replace('d', ''))
                years = days / 365.0
            else:
                years = 1.0  # Par défaut 1 an
            
            rentabilite_annuelle_pct_list = []
            for d in domains:
                trades = stats['by_domain'][d]['trades']
                gain = stats['by_domain'][d]['gain']
                nb_symbols = symbols_per_domain.get(d, 1)
                
                # Capital investi: 50€ par stock
                capital_investi = nb_symbols * 50.0
                
                # Rendement total en %
                rendement_total_pct = (gain / capital_investi) * 100 if capital_investi > 0 else 0.0
                
                # Annualiser le rendement
                rendement_annuel_pct = rendement_total_pct / years if years > 0 else rendement_total_pct
                
                rentabilite_annuelle_pct_list.append(rendement_annuel_pct)
            
            # Créer figure avec moins de graphiques pour plus de clarté (2 lignes x 2 colonnes)
            fig = Figure(figsize=(16, 10), dpi=90)
            
            # Subplot 1: Bar chart taux de réussite
            ax1 = fig.add_subplot(2, 2, 1)
            colors = ['green' if t >= 60 else 'orange' if t >= 40 else 'red' for t in taux_list]
            bars1 = ax1.bar(domains, taux_list, color=colors, alpha=0.75, edgecolor='black', linewidth=1.2)
            ax1.set_ylabel('Taux de réussite (%)', fontweight='bold', fontsize=11)
            ax1.set_title('Taux de réussite', fontweight='bold', fontsize=13, pad=10)
            ax1.set_ylim(0, 105)
            ax1.axhline(y=50, color='gray', linestyle='--', alpha=0.5, linewidth=1.5)
            ax1.tick_params(axis='x', rotation=45, labelsize=9)
            ax1.tick_params(axis='y', labelsize=9)
            ax1.grid(axis='y', alpha=0.3, linestyle='--')
            for i, v in enumerate(taux_list):
                ax1.text(i, v + 3, f'{v:.1f}%', ha='center', fontsize=10, fontweight='bold')
            
            # Subplot 2: Bar chart gain total
            ax2 = fig.add_subplot(2, 2, 2)
            colors_gain = ['green' if g > 0 else 'red' for g in gain_list]
            bars2 = ax2.bar(domains, gain_list, color=colors_gain, alpha=0.75, edgecolor='black', linewidth=1.2)
            ax2.set_ylabel('Gain total ($)', fontweight='bold', fontsize=11)
            ax2.set_title('Gain brut total', fontweight='bold', fontsize=13, pad=10)
            ax2.axhline(y=0, color='black', linewidth=1.5)
            ax2.tick_params(axis='x', rotation=45, labelsize=9)
            ax2.tick_params(axis='y', labelsize=9)
            ax2.grid(axis='y', alpha=0.3, linestyle='--')
            for i, v in enumerate(gain_list):
                offset = 20 if v > 0 else -40
                ax2.text(i, v + offset, f'${v:.0f}', ha='center', fontsize=10, fontweight='bold')
            
            # Subplot 3: Bar chart rentabilité annuelle en %
            ax3 = fig.add_subplot(2, 2, 3)
            colors_rentabilite = ['darkgreen' if r > 50 else 'green' if r > 10 else 'orange' if r > 0 else 'red' for r in rentabilite_annuelle_pct_list]
            bars3 = ax3.bar(domains, rentabilite_annuelle_pct_list, color=colors_rentabilite, alpha=0.75, edgecolor='black', linewidth=1.2)
            ax3.set_ylabel('Rentabilité annuelle (%)', fontweight='bold', fontsize=11)
            ax3.set_title('Rentabilité annualisée', fontweight='bold', fontsize=13, pad=10)
            ax3.axhline(y=0, color='black', linewidth=1.5)
            ax3.axhline(y=10, color='green', linestyle='--', alpha=0.4, linewidth=1)
            ax3.tick_params(axis='x', rotation=45, labelsize=9)
            ax3.tick_params(axis='y', labelsize=9)
            ax3.grid(axis='y', alpha=0.3, linestyle='--')
            for i, v in enumerate(rentabilite_annuelle_pct_list):
                offset = max(abs(v) * 0.1, 3) if v > 0 else -max(abs(v) * 0.1, 5)
                ax3.text(i, v + offset, f'{v:.1f}%', ha='center', fontsize=10, fontweight='bold')
            
            # Subplot 4: Distribution trades + symboles combinée
            ax4 = fig.add_subplot(2, 2, 4)
            x_pos = range(len(domains))
            width = 0.35
            
            # Normaliser pour affichage sur même échelle (éviter division par zéro)
            max_trades = max(trades_list) if trades_list and max(trades_list) > 0 else 1
            max_symbols = max(symbols_list) if symbols_list and max(symbols_list) > 0 else 1
            trades_normalized = [t / max_trades * 100 if max_trades > 0 else 0 for t in trades_list]
            symbols_normalized = [s / max_symbols * 100 if max_symbols > 0 else 0 for s in symbols_list]
            
            bars4a = ax4.bar([x - width/2 for x in x_pos], trades_normalized, width, 
                            label='Trades', color='steelblue', alpha=0.75, edgecolor='black', linewidth=1)
            bars4b = ax4.bar([x + width/2 for x in x_pos], symbols_normalized, width,
                            label='Symboles', color='coral', alpha=0.75, edgecolor='black', linewidth=1)
            
            ax4.set_ylabel('Distribution (normalisée)', fontweight='bold', fontsize=11)
            ax4.set_title('Trades et Symboles par secteur', fontweight='bold', fontsize=13, pad=10)
            ax4.set_xticks(x_pos)
            ax4.set_xticklabels(domains, rotation=45, ha='right', fontsize=9)
            ax4.tick_params(axis='y', labelsize=9)
            ax4.legend(loc='upper right', fontsize=10)
            ax4.grid(axis='y', alpha=0.3, linestyle='--')
            
            # Ajouter les valeurs réelles au-dessus des barres
            for i, (t, s) in enumerate(zip(trades_list, symbols_list)):
                ax4.text(i - width/2, trades_normalized[i] + 3, str(int(t)), 
                        ha='center', fontsize=8, fontweight='bold')
                ax4.text(i + width/2, symbols_normalized[i] + 3, str(int(s)), 
                        ha='center', fontsize=8, fontweight='bold')
            
            fig.tight_layout()
            canvas = FigureCanvas(fig)
            
            # Ajouter à l'onglet Graphiques
            self.charts_scroll_layout.addWidget(title)
            self.charts_scroll_layout.addWidget(summary_label)
            self.charts_scroll_layout.addWidget(canvas)
            self.charts_scroll_layout.addStretch()
            
        except Exception as e:
            print(f"❌ Erreur populate_charts_tab: {e}")
            import traceback
            traceback.print_exc()
    
    def populate_comparisons_tab(self):
        """Onglet Comparaisons: permet de sélectionner jusqu'à 100 symboles et les comparer."""
        try:
            # Nettoyer l'onglet Comparaisons
            while self.comparisons_layout.count() > 0:
                widget = self.comparisons_layout.takeAt(0).widget()
                if widget:
                    widget.deleteLater()
            
            # Titre
            title = QLabel("📊 Comparaison personnalisée de symboles (max 100)")
            title.setStyleSheet("font-weight: bold; font-size: 12px;")
            self.comparisons_layout.addWidget(title)
            
            # Récupérer tous les symboles disponibles
            all_symbols = []
            for row in range(self.merged_table.rowCount()):
                try:
                    sym = self.merged_table.item(row, 0).text()
                    if sym:
                        all_symbols.append(sym)
                except Exception:
                    continue
            
            if not all_symbols:
                self.comparisons_layout.addWidget(QLabel("Aucun symbole disponible pour la comparaison."))
                return
            
            # Conteneur de sélection avec scroll
            selection_container = QWidget()
            selection_layout = QVBoxLayout(selection_container)
            
            # Label pour sélection avec boutons rapides
            select_header_layout = QHBoxLayout()
            select_label = QLabel("✓ Sélectionnez jusqu'à 100 symboles:")
            select_label.setStyleSheet("font-weight: bold; font-size: 10px;")
            select_header_layout.addWidget(select_label)
            select_header_layout.addStretch()
            
            # ✅ Boutons de sélection rapide
            select_all_btn = QPushButton("Tout sélectionner")
            select_all_btn.setMaximumWidth(120)
            select_all_btn.setStyleSheet("background-color: #2196F3; color: white; font-size: 9px; padding: 4px;")
            
            deselect_all_btn = QPushButton("Tout désélectionner")
            deselect_all_btn.setMaximumWidth(130)
            deselect_all_btn.setStyleSheet("background-color: #757575; color: white; font-size: 9px; padding: 4px;")
            
            select_header_layout.addWidget(select_all_btn)
            select_header_layout.addWidget(deselect_all_btn)
            
            select_label_widget = QWidget()
            select_label_widget.setLayout(select_header_layout)
            selection_layout.addWidget(select_label_widget)
            
            # Sélecteur de date pour comparaison historique
            date_container = QWidget()
            date_layout = QHBoxLayout(date_container)
            date_label = QLabel("📅 Date de référence (optionnel):")
            date_label.setStyleSheet("font-size: 10px;")
            
            from PyQt5.QtWidgets import QDateEdit
            from PyQt5.QtCore import QDate
            self.comparison_date_edit = QDateEdit()
            self.comparison_date_edit.setCalendarPopup(True)
            self.comparison_date_edit.setDate(QDate.currentDate())
            self.comparison_date_edit.setMaximumDate(QDate.currentDate())
            self.comparison_date_edit.setMinimumDate(QDate(2020, 1, 1))
            self.comparison_date_edit.setDisplayFormat("dd/MM/yyyy")
            
            self.use_historical_check = QCheckBox("Comparer avec données historiques")
            self.use_historical_check.setToolTip("Analyser les symboles à la date sélectionnée et voir l'évolution réelle depuis")
            
            date_layout.addWidget(date_label)
            date_layout.addWidget(self.comparison_date_edit)
            date_layout.addWidget(self.use_historical_check)
            date_layout.addStretch()
            selection_layout.addWidget(date_container)
            
            # Scroll area pour les checkboxes
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setMaximumHeight(120)
            checkbox_container = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_container)
            
            # Créer les checkboxes
            checkboxes = {}
            for sym in sorted(all_symbols):
                cb = QCheckBox(sym)
                checkboxes[sym] = cb
                checkbox_layout.addWidget(cb)
            
            scroll.setWidget(checkbox_container)
            selection_layout.addWidget(scroll)
            
            # ✅ Connexions des boutons de sélection rapide
            def on_select_all():
                for cb in checkboxes.values():
                    cb.setChecked(True)
            
            def on_deselect_all():
                for cb in checkboxes.values():
                    cb.setChecked(False)
            
            select_all_btn.clicked.connect(on_select_all)
            deselect_all_btn.clicked.connect(on_deselect_all)
            
            # Boutons d'action
            button_layout = QHBoxLayout()
            compare_btn = QPushButton("Comparer les symboles sélectionnés")
            compare_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
            reset_btn = QPushButton("Réinitialiser")
            reset_btn.setStyleSheet("background-color: #f44336; color: white;")
            
            button_layout.addWidget(compare_btn)
            button_layout.addWidget(reset_btn)
            selection_layout.addLayout(button_layout)
            
            self.comparisons_layout.addWidget(selection_container)
            
            # Zone de résultats de comparaison (initialement vide)
            results_scroll = QScrollArea()
            results_scroll.setWidgetResizable(True)
            self.comparison_results = QWidget()
            self.comparison_results_layout = QVBoxLayout(self.comparison_results)
            results_scroll.setWidget(self.comparison_results)
            
            self.comparisons_layout.addWidget(results_scroll)
            self.comparisons_layout.addStretch()
            
            # Connexions des boutons
            def on_compare():
                selected = [sym for sym, cb in checkboxes.items() if cb.isChecked()]
                if not selected:
                    QMessageBox.warning(self, "Erreur", "Sélectionnez au moins 1 symbole pour comparer")
                    return
                if len(selected) > 100:
                    QMessageBox.warning(self, "Erreur", "Maximum 100 symboles à la fois")
                    return
                
                # Nettoyer les résultats précédents
                while self.comparison_results_layout.count() > 0:
                    w = self.comparison_results_layout.takeAt(0).widget()
                    if w:
                        w.deleteLater()
                
                # Check if historical comparison requested
                use_historical = self.use_historical_check.isChecked()
                if use_historical:
                    selected_date = self.comparison_date_edit.date()
                    historical_date_str = selected_date.toString("yyyy-MM-dd")
                    self._generate_historical_comparison_table(selected, historical_date_str)
                else:
                    # Générer le tableau comparatif
                    self._generate_comparison_table(selected)
            
            def on_reset():
                for cb in checkboxes.values():
                    cb.setChecked(False)
                while self.comparisons_layout.count() > 0:
                    w = self.comparisons_layout.takeAt(0).widget()
                    if w:
                        w.deleteLater()
            
            compare_btn.clicked.connect(on_compare)
            reset_btn.clicked.connect(on_reset)
            
        except Exception as e:
            print(f"❌ Erreur populate_comparisons_tab: {e}")
            import traceback
            traceback.print_exc()
    
    def _generate_comparison_table(self, symbols_to_compare):
        """Génère un tableau comparatif pour les symboles sélectionnés avec classement par pertinence."""
        try:
            def safe_float(value, default=0.0):
                try:
                    if value is None:
                        return default
                    return float(str(value).replace('%', '').replace('$', '').replace(',', '').strip())
                except Exception:
                    return default

            def table_text(row, col, default='0'):
                item = self.merged_table.item(row, col)
                return item.text() if item else default

            def clone_table_item(source_item, fallback_text=''):
                text = source_item.text() if source_item else fallback_text

                def _format_numeric_like_source(display_text, numeric_value):
                    """Format all numeric display values with at most 2 decimals in comparison table."""
                    txt = str(display_text or '')
                    if numeric_value is None:
                        return txt
                    if '$' in txt:
                        return f"${float(numeric_value):.2f}"
                    if '%' in txt:
                        return f"{float(numeric_value):.2f}%"
                    if txt.strip().endswith('x'):
                        return f"{float(numeric_value):.2f}x"

                    clean_txt = txt.strip().replace(',', '')
                    if '.' in clean_txt or 'e' in clean_txt.lower():
                        return f"{float(numeric_value):.2f}"
                    return txt

                if source_item:
                    edit_value = source_item.data(Qt.EditRole)
                    if isinstance(edit_value, (int, float)):
                        text = _format_numeric_like_source(text, edit_value)

                item = QTableWidgetItem(text)
                if source_item:
                    # Keep the already-formatted text from source_item to preserve UI rounding.
                    item.setData(Qt.EditRole, source_item.data(Qt.EditRole))
                    item.setForeground(source_item.foreground())
                    item.setBackground(source_item.background())
                    item.setFont(source_item.font())
                    item.setTextAlignment(source_item.textAlignment())
                    item.setFlags(source_item.flags())
                    tooltip = source_item.toolTip()
                    if tooltip:
                        item.setToolTip(tooltip)
                return item

            # Récupérer les données pour chaque symbole en recopiant les colonnes du tableau résultats
            symbols_data = {}
            for row in range(self.merged_table.rowCount()):
                try:
                    sym_item = self.merged_table.item(row, 0)
                    if not sym_item:
                        continue
                    sym = sym_item.text()
                    if sym in symbols_to_compare:
                        score = safe_float(table_text(row, 2))
                        prix = safe_float(table_text(row, 3))
                        rsi = safe_float(table_text(row, 5))
                        domaine = table_text(row, 7, 'N/A')
                        score_seuil = safe_float(table_text(row, 9))
                        fiab = safe_float(table_text(row, 10))
                        trades = int(safe_float(table_text(row, 11)))
                        gagnants = int(safe_float(table_text(row, 12)))
                        rev_growth = safe_float(table_text(row, 13))
                        ebitda = safe_float(table_text(row, 14))
                        fcf = safe_float(table_text(row, 15))
                        debt_to_equity = safe_float(table_text(row, 16))
                        market_cap = safe_float(table_text(row, 17))
                        roe = safe_float(table_text(row, 18))
                        dprice = safe_float(table_text(row, 19))
                        var5j = safe_float(table_text(row, 20))
                        drsi = safe_float(table_text(row, 21))
                        dvol = safe_float(table_text(row, 22))
                        gain_total = safe_float(table_text(row, 23))
                        gain_moyen = safe_float(table_text(row, 24))
                        consensus = table_text(row, 25, 'N/A')
                        
                        symbols_data[sym] = {
                            'Score': score,
                            'Prix': prix,
                            'RSI': rsi,
                            'Domaine': domaine,
                            'Score/Seuil': score_seuil,
                            'Fiabilité (%)': fiab,
                            'Nb Trades': trades,
                            'Gagnants': gagnants,
                            'Rev. Growth (%)': rev_growth,
                            'EBITDA Yield (%)': ebitda,
                            'FCF Yield (%)': fcf,
                            'D/E Ratio': debt_to_equity,
                            'Market Cap (B$)': market_cap,
                            'ROE (%)': roe,
                            'dPrice': dprice,
                            'Var5j (%)': var5j,
                            'dRSI': drsi,
                            'dVolRel': dvol,
                            'Gain total ($)': gain_total,
                            'Gain moyen ($)': gain_moyen,
                            'Consensus': consensus
                        }
                except Exception:
                    continue

            if not symbols_data:
                self.comparison_results_layout.addWidget(QLabel("Aucune donnée exploitable pour la comparaison."))
                return
            
            # Logique multicritere par rang:
            # pour chaque critere, le meilleur gagne n points, puis n-1 ... jusqu'a 1.
            # Pertinence (%) = points_total / (n * m) * 100
            # (n = nombre de stocks compares, m = nombre de criteres)
            criteria_config = [
                {'key': 'Score', 'label': 'Score', 'order': 'desc'},
                {'key': 'Score/Seuil', 'label': 'Score/Seuil', 'order': 'desc'},
                {'key': 'Fiabilité (%)', 'label': 'Fiabilite (%)', 'order': 'desc'},
                {'key': 'Nb Trades', 'label': 'Nb Trades', 'order': 'desc'},
                {'key': 'Gagnants', 'label': 'Gagnants', 'order': 'desc'},
                {'key': 'Rev. Growth (%)', 'label': 'Rev Growth (%)', 'order': 'desc'},
                {'key': 'EBITDA Yield (%)', 'label': 'EBITDA (%)', 'order': 'desc'},
                {'key': 'FCF Yield (%)', 'label': 'FCF (%)', 'order': 'desc'},
                {'key': 'D/E Ratio', 'label': 'D/E', 'order': 'asc'},
                {'key': 'Market Cap (B$)', 'label': 'Market Cap (B$)', 'order': 'desc'},
                {'key': 'ROE (%)', 'label': 'ROE (%)', 'order': 'desc'},
                {'key': 'dPrice', 'label': 'dPrice', 'order': 'desc'},
                {'key': 'Var5j (%)', 'label': 'Var5j (%)', 'order': 'asc'},
                {'key': 'dRSI', 'label': 'dRSI', 'order': 'asc'},
                {'key': 'dVolRel', 'label': 'dVolRel', 'order': 'desc'},
                {'key': 'Gain total ($)', 'label': 'Gain total ($)', 'order': 'desc'},
                {'key': 'Gain moyen ($)', 'label': 'Gain moyen ($)', 'order': 'desc'},
            ]

            selected_symbols = [sym for sym in symbols_to_compare if sym in symbols_data]
            n_stocks = len(selected_symbols)
            m_criteria = len(criteria_config)
            max_points = n_stocks * m_criteria if n_stocks > 0 and m_criteria > 0 else 1

            points_by_symbol = {sym: 0 for sym in selected_symbols}

            for criterion in criteria_config:
                key = criterion['key']
                reverse = criterion['order'] == 'desc'
                ranked = sorted(
                    selected_symbols,
                    key=lambda s: safe_float(symbols_data[s].get(key, 0.0), 0.0),
                    reverse=reverse,
                )
                for rank_idx, sym in enumerate(ranked):
                    # n_stocks points pour le 1er, ... 1 point pour le dernier
                    points_by_symbol[sym] += (n_stocks - rank_idx)

            pertinence_scores = {
                sym: (points_by_symbol[sym] / max_points) * 100.0
                for sym in selected_symbols
            }
            
            # Classer par pertinence (décroissant), puis garde-fous de tri secondaires
            sorted_symbols = sorted(
                selected_symbols,
                key=lambda x: (pertinence_scores[x], symbols_data[x]['Score/Seuil'], symbols_data[x]['Score']),
                reverse=True
            )
            
            # Créer un tableau QTableWidget pour afficher la comparaison
            table = QTableWidget()
            columns = ['Rang', 'Symbole', 'Signal', 'Score', 'Prix', 'Tendance', 'RSI', 'Volume moyen($)', 'Domaine', 'Cap Range',
                      'Score/Seuil', 'Fiabilité (%)', 'Nb Trades', 'Gagnants', 'Rev Growth (%)', 'EBITDA (%)', 'FCF (%)',
                      'D/E', 'Market Cap (B$)', 'ROE (%)', 'dPrice', 'Var5j (%)', 'dRSI', 'dVolRel', 'Gain total ($)',
                      'Gain moyen ($)', 'Consensus', 'Pertinence']
            table.setColumnCount(len(columns))
            table.setHorizontalHeaderLabels(columns)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            
            # Remplir le tableau
            for rank, sym in enumerate(sorted_symbols, 1):
                data = symbols_data[sym]
                pertinence = pertinence_scores[sym]
                
                row = table.rowCount()
                table.insertRow(row)
                
                # Rang
                item = QTableWidgetItem(str(rank))
                if rank == 1:
                    item.setBackground(QColor(144, 238, 144))  # Vert pour 1er
                elif rank == 2:
                    item.setBackground(QColor(211, 211, 211))  # Gris pour 2ème
                table.setItem(row, 0, item)

                # Rechercher la ligne source une seule fois pour recopier toutes les colonnes
                source_row = None
                for r in range(self.merged_table.rowCount()):
                    item = self.merged_table.item(r, 0)
                    if item and item.text() == sym:
                        source_row = r
                        break
                if source_row is None:
                    continue

                # Copier les cellules 0..25 depuis le tableau source vers 1..26 (0 est reserve au rang)
                for source_col, target_col in enumerate(range(1, 27), start=0):
                    source_item = self.merged_table.item(source_row, source_col)
                    table.setItem(row, target_col, clone_table_item(source_item, table_text(source_row, source_col, '')))

                # Pertinence affichée telle quelle pour garder le tri du tableau source
                item = QTableWidgetItem(f"{pertinence:.2f}%")
                item.setData(Qt.EditRole, pertinence)
                item.setBackground(Qt.yellow)
                item.setToolTip(f"Points: {points_by_symbol.get(sym, 0)} / {max_points}")
                table.setItem(row, 27, item)
            
            table.setSortingEnabled(True)
            table.setMinimumHeight(300)
            
            # Ajouter un résumé
            summary = QLabel(
                f"📊 Comparaison de {len(sorted_symbols)} symbole(s) | "
                f"🧮 Méthode: rang multicritère ({m_criteria} critères) | "
                f"🥇 Meilleur: {sorted_symbols[0]} "
                f"(Pertinence: {pertinence_scores[sorted_symbols[0]]:.1f}%)"
            )
            summary.setStyleSheet("background-color: #e3f2fd; padding: 6px; border-radius: 4px; font-weight: bold;")

            sens_parts = []
            for c in criteria_config:
                arrow = '↘ décroissant' if c['order'] == 'desc' else '↗ croissant'
                sens_parts.append(f"{c['label']}: {arrow}")
            senses_label = QLabel("🧭 Sens des critères: " + " | ".join(sens_parts))
            senses_label.setWordWrap(True)
            senses_label.setStyleSheet("background-color: #f6f8fa; padding: 6px; border-radius: 4px; font-size: 9px;")
            
            self.comparison_results_layout.addWidget(summary)
            self.comparison_results_layout.addWidget(senses_label)
            self.comparisons_layout.addWidget(table)
            
        except Exception as e:
            print(f"❌ Erreur _generate_comparison_table: {e}")
            import traceback
            traceback.print_exc()
            error_label = QLabel(f"Erreur: {e}")
            self.comparison_results_layout.addWidget(error_label)
    
    def _generate_historical_comparison_table(self, symbols_to_compare, historical_date):
        """
        Génère un tableau de comparaison historique avec analyse complète.
        Télécharge intelligemment 18 mois de données pour le backtest annuel.
        Évite les retéléchargements en utilisant un cache intelligent.
        """
        try:
            from datetime import datetime, timedelta
            import pandas as pd
            import yfinance as yf
            from pathlib import Path
            
            # Créer un répertoire pour le cache historique
            cache_dir = Path("data_cache/historical")
            cache_dir.mkdir(parents=True, exist_ok=True)
            
            # Afficher le chargement
            loading_label = QLabel(f"⏳ Téléchargement intelligent des données (18 mois) pour {historical_date}...")
            loading_label.setStyleSheet("font-size: 11px; color: blue; padding: 10px;")
            self.comparison_results_layout.addWidget(loading_label)
            QApplication.processEvents()
            
            # Convertir la date
            target_date = datetime.strptime(historical_date, "%Y-%m-%d")
            today = datetime.now()
            
            # Vérifier que c'est une date passée
            if target_date >= today:
                QMessageBox.warning(self, "Erreur", "Veuillez sélectionner une date passée")
                loading_label.deleteLater()
                return
            
            # Période à télécharger : 18 mois avant la date cible
            # (12 mois de backtest + 6 mois pour les indicateurs)
            dl_start_date = target_date - timedelta(days=550)  # ~18 mois
            dl_end_date = target_date + timedelta(days=1)  # Inclure la date cible
            
            # Stocker les données historiques et actuelles
            historical_data = {}
            actual_performance = {}
            
            # Fonction interne pour gérer le cache intelligent
            def get_or_download_data(symbol, start, end):
                """Récupère du cache ou télécharge avec gestion intelligente"""
                cache_file = cache_dir / f"{symbol}_hist.pkl"
                
                # Vérifier si le cache existe et est suffisamment complet
                if cache_file.exists():
                    try:
                        cached_df = pd.read_pickle(cache_file)
                        cached_meta = cache_dir / f"{symbol}_hist_meta.txt"
                        
                        # Vérifier les metadata (période couverte)
                        if cached_meta.exists():
                            with open(cached_meta, 'r') as f:
                                meta = f.read().strip()
                                cached_start, cached_end = meta.split('|')
                                cached_start = datetime.strptime(cached_start, "%Y-%m-%d")
                                cached_end = datetime.strptime(cached_end, "%Y-%m-%d")
                                
                                # Si le cache couvre la période requise
                                if cached_start <= start and cached_end >= end:
                                    print(f"✓ Cache valide pour {symbol}")
                                    return cached_df
                    except Exception as e:
                        print(f"⚠️ Erreur lecture cache {symbol}: {e}")
                
                # Télécharger les données manquantes
                print(f"📥 Téléchargement {symbol} ({start.strftime('%Y-%m-%d')} à {end.strftime('%Y-%m-%d')})")
                try:
                    df = yf.download(
                        symbol, 
                        start=start.strftime("%Y-%m-%d"), 
                        end=end.strftime("%Y-%m-%d"), 
                        progress=False,
                        timeout=30
                    )
                    
                    if not df.empty:
                        # Sauvegarder en cache avec metadata
                        df.to_pickle(cache_file)
                        with open(cache_dir / f"{symbol}_hist_meta.txt", 'w') as f:
                            f.write(f"{start.strftime('%Y-%m-%d')}|{end.strftime('%Y-%m-%d')}")
                        print(f"💾 Cache sauvegardé pour {symbol}")
                        return df
                except Exception as e:
                    print(f"❌ Erreur téléchargement {symbol}: {e}")
                    return None
                
                return None
            
            # Télécharger les données pour tous les symboles
            for idx, symbol in enumerate(symbols_to_compare):
                try:
                    # Mettre à jour le label de progression
                    loading_label.setText(
                        f"⏳ Traitement {symbol} ({idx+1}/{len(symbols_to_compare)}) - "
                        f"Téléchargement intelligent..."
                    )
                    QApplication.processEvents()
                    
                    # Récupérer/télécharger les données
                    df = get_or_download_data(symbol, dl_start_date, dl_end_date)
                    
                    if df is None or df.empty:
                        print(f"⚠️ Pas de données pour {symbol}")
                        continue

                    # Uniformiser tous les calculs de prix en USD.
                    try:
                        df = qsi._normalize_prices_to_usd(symbol, df)
                    except Exception:
                        pass
                    
                    # Normaliser l'index
                    if df.index.name is None or df.index.name != 'Date':
                        df = df.reset_index()
                        if 'Date' in df.columns:
                            df['Date'] = pd.to_datetime(df['Date'])
                            df.set_index('Date', inplace=True)
                    
                    df.index = df.index.tz_localize(None) if df.index.tz is not None else df.index
                    
                    # Trouver le prix à la date cible (ou le plus proche)
                    closest_date = df.index[df.index <= target_date].max() if any(df.index <= target_date) else None
                    
                    if closest_date is None:
                        print(f"⚠️ Pas de données pour {symbol} avant {historical_date}")
                        continue
                    
                    # Données à la date historique et aujourd'hui
                    historical_price = df.loc[closest_date]['Close']
                    current_price = df.iloc[-1]['Close']
                    
                    # Calculer la performance réelle (%)
                    performance_pct = ((current_price - historical_price) / historical_price) * 100
                    
                    # Historique jusqu'à la date cible
                    historical_idx = df.index.get_loc(closest_date)
                    hist_df = df.iloc[:historical_idx+1].copy()
                    
                    # === CALCUL DES INDICATEURS À LA DATE HISTORIQUE ===
                    
                    # 1. RSI (14 jours)
                    hist_rsi = self._calculate_rsi(hist_df['Close'], period=14)
                    
                    # 2. MACD
                    hist_macd = self._calculate_macd(hist_df['Close'])
                    
                    # 3. Bandes de Bollinger (20 jours)
                    hist_bb = self._calculate_bollinger_bands(hist_df['Close'], period=20)
                    
                    # 4. Volume Relatif (20 jours)
                    if len(hist_df) >= 20:
                        avg_vol = hist_df['Volume'].rolling(window=20).mean().iloc[-1]
                        current_vol = hist_df['Volume'].iloc[-1]
                        vol_rel = (current_vol / avg_vol) if avg_vol > 0 else 1.0
                    else:
                        vol_rel = 1.0
                    
                    # 5. Tendance de prix (30 jours)
                    if len(hist_df) >= 30:
                        price_trend = hist_df['Close'].iloc[-30:].pct_change().mean() * 100
                    else:
                        price_trend = 0.0
                    
                    # 6. Volatilité (20 jours)
                    if len(hist_df) >= 20:
                        volatility = hist_df['Close'].pct_change().rolling(window=20).std().iloc[-1] * 100
                    else:
                        volatility = 0.0
                    
                    # === CALCUL DU SCORE DE PERTINENCE HISTORIQUE ===
                    
                    # Score basé sur les indicateurs
                    rsi_score = 0
                    if hist_rsi < 30:
                        rsi_score = 25  # Survendu (bon signal d'achat)
                    elif hist_rsi > 70:
                        rsi_score = 0   # Suracheté (mauvais)
                    else:
                        rsi_score = 15  # Neutre
                    
                    # Score MACD
                    macd_score = 15 if hist_macd > 0 else 0
                    
                    # Score Bollinger Bands
                    bb_score = 10 if hist_bb else 0
                    
                    # Score Volume
                    vol_score = 10 if vol_rel > 1.2 else 5
                    
                    # Score Tendance
                    trend_score = 15 if price_trend > 0 else 5
                    
                    # Score Volatilité (haute volatilité = plus d'opportunité)
                    vol_std_score = 10 if volatility > 2.0 else 5
                    
                    # Score basé sur performance réelle (validation)
                    actual_score = min(25, max(0, performance_pct / 10))  # Max 25 points
                    
                    # Total
                    pertinence_score = (rsi_score + macd_score + bb_score + 
                                       vol_score + trend_score + vol_std_score + actual_score)
                    
                    historical_data[symbol] = {
                        'Date': closest_date.strftime("%Y-%m-%d"),
                        'Prix Historique': float(historical_price),
                        'Prix Actuel': float(current_price),
                        'Performance (%)': float(performance_pct),
                        'RSI': float(hist_rsi),
                        'MACD': float(hist_macd),
                        'Volume Rel': float(vol_rel),
                        'Tendance (%)': float(price_trend),
                        'Volatilité (%)': float(volatility),
                        'Pertinence': float(pertinence_score)
                    }
                    
                    actual_performance[symbol] = performance_pct
                    
                except Exception as e:
                    print(f"❌ Erreur pour {symbol}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # Supprimer le label de chargement
            loading_label.deleteLater()
            
            if not historical_data:
                error_label = QLabel("❌ Aucune donnée historique disponible pour les symboles sélectionnés")
                error_label.setStyleSheet("color: red; padding: 10px;")
                self.comparison_results_layout.addWidget(error_label)
                return
            
            # Trier par performance réelle (meilleure performance = meilleur)
            sorted_symbols = sorted(historical_data.keys(), key=lambda x: actual_performance[x], reverse=True)
            
            # Créer le tableau
            table = QTableWidget()
            table.setColumnCount(11)
            table.setRowCount(len(sorted_symbols))
            table.setHorizontalHeaderLabels([
                'Rang', 'Symbole', 'Date', 'Prix Historique', 'Prix Actuel',
                'Performance (%)', 'RSI', 'MACD', 'Vol. Rel', 'Volatilité (%)', 'Avis'
            ])
            
            for row, symbol in enumerate(sorted_symbols):
                data = historical_data[symbol]
                
                # Rang
                item = QTableWidgetItem(str(row + 1))
                if row == 0:
                    item.setBackground(QColor(144, 238, 144))  # Vert pour 1er
                elif row == 1:
                    item.setBackground(QColor(211, 211, 211))  # Gris pour 2ème
                table.setItem(row, 0, item)
                
                # Symbole
                table.setItem(row, 1, QTableWidgetItem(symbol))
                
                # Date d'analyse
                table.setItem(row, 2, QTableWidgetItem(data['Date']))
                
                # Prix historique
                item = QTableWidgetItem(f"${data['Prix Historique']:.2f}")
                item.setData(Qt.EditRole, data['Prix Historique'])
                table.setItem(row, 3, item)
                
                # Prix actuel
                item = QTableWidgetItem(f"${data['Prix Actuel']:.2f}")
                item.setData(Qt.EditRole, data['Prix Actuel'])
                table.setItem(row, 4, item)
                
                # Performance
                perf = data['Performance (%)']
                item = QTableWidgetItem(f"{perf:+.2f}%")
                item.setData(Qt.EditRole, perf)
                if perf > 0:
                    item.setForeground(QColor(0, 128, 0))  # Vert
                else:
                    item.setForeground(QColor(255, 0, 0))  # Rouge
                table.setItem(row, 5, item)
                
                # RSI
                rsi = data['RSI']
                item = QTableWidgetItem(f"{rsi:.2f}")
                item.setData(Qt.EditRole, rsi)
                if rsi < 30 or rsi > 70:
                    item.setForeground(QColor(255, 140, 0))  # Orange (extrême)
                table.setItem(row, 6, item)
                
                # MACD
                macd = data['MACD']
                item = QTableWidgetItem(f"{macd:+.4f}")
                item.setData(Qt.EditRole, macd)
                table.setItem(row, 7, item)
                
                # Volume Relatif
                vol = data['Volume Rel']
                item = QTableWidgetItem(f"{vol:.2f}x")
                item.setData(Qt.EditRole, vol)
                table.setItem(row, 8, item)
                
                # Volatilité
                vol_std = data['Volatilité (%)']
                item = QTableWidgetItem(f"{vol_std:.2f}%")
                item.setData(Qt.EditRole, vol_std)
                table.setItem(row, 9, item)
                
                # Avis (justification)
                avis = self._generate_historical_verdict(data)
                table.setItem(row, 10, QTableWidgetItem(avis))
            
            table.setSortingEnabled(True)
            table.setMinimumHeight(350)
            table.resizeColumnsToContents()
            
            # Résumé et statistiques
            best_symbol = sorted_symbols[0]
            best_perf = actual_performance[best_symbol]
            worst_symbol = sorted_symbols[-1]
            worst_perf = actual_performance[worst_symbol]
            avg_perf = sum(actual_performance.values()) / len(actual_performance)
            winners = sum(1 for p in actual_performance.values() if p > 0)
            
            summary = QLabel(
                f"📈 Historique ({historical_date}) | "
                f"🥇 {best_symbol} ({best_perf:+.1f}%) | "
                f"📊 Moyenne: {avg_perf:+.1f}% | "
                f"✓ {winners}/{len(actual_performance)} gagnants"
            )
            summary.setStyleSheet("background-color: #fff9c4; padding: 8px; border-radius: 4px; font-weight: bold;")
            
            info_label = QLabel(
                f"💡 Analyse complète : 18 mois téléchargés intelligemment (12 mois backtest + 6 mois indicateurs). "
                f"Le cache est utilisé pour éviter les retéléchargements. "
                f"Les symboles sont classés par performance réelle depuis {historical_date}."
            )
            info_label.setWordWrap(True)
            info_label.setStyleSheet("background-color: #e1f5fe; padding: 6px; border-radius: 4px; font-size: 9px;")
            
            self.comparison_results_layout.addWidget(summary)
            self.comparisons_layout.addWidget(info_label)
            self.comparisons_layout.addWidget(table)
            
        except Exception as e:
            print(f"❌ Erreur _generate_historical_comparison_table: {e}")
            import traceback
            traceback.print_exc()
            error_label = QLabel(f"Erreur: {e}")
            self.comparisons_layout.addWidget(error_label)
    
    def _calculate_rsi(self, prices, period=14):
        """Calcule le RSI (Relative Strength Index)"""
        delta = prices.diff()
        gain = delta.where(delta > 0, 0).rolling(window=period).mean()
        loss = -delta.where(delta < 0, 0).rolling(window=period).mean()
        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return float(rsi.iloc[-1]) if not rsi.empty else 50.0
    
    def _calculate_macd(self, prices, fast=12, slow=26):
        """Calcule le MACD (Moving Average Convergence Divergence)"""
        ema_fast = prices.ewm(span=fast).mean()
        ema_slow = prices.ewm(span=slow).mean()
        macd = ema_fast - ema_slow
        return float(macd.iloc[-1]) if not macd.empty else 0.0
    
    def _calculate_bollinger_bands(self, prices, period=20, num_std=2):
        """Calcule les Bandes de Bollinger et retourne si le prix est en extrême"""
        sma = prices.rolling(window=period).mean()
        std = prices.rolling(window=period).std()
        upper_band = sma + (std * num_std)
        lower_band = sma - (std * num_std)
        current_price = prices.iloc[-1]
        # Retourne True si le prix touche un extrême
        return current_price >= upper_band.iloc[-1] or current_price <= lower_band.iloc[-1]
    
    def _generate_historical_verdict(self, data):
        """Génère un avis pointu basé sur les indicateurs"""
        avis_parts = []
        
        rsi = data['RSI']
        if rsi < 30:
            avis_parts.append("↓ Survendu")
        elif rsi > 70:
            avis_parts.append("↑ Suracheté")
        
        if data['MACD'] > 0:
            avis_parts.append("▲ MACD+")
        else:
            avis_parts.append("▼ MACD-")
        
        if data['Volume Rel'] > 1.5:
            avis_parts.append("📈 Vol▲")
        
        perf = data['Performance (%)']
        if perf > 20:
            avis_parts.append("🚀 +20%")
        elif perf < -10:
            avis_parts.append("📉 -10%")
        
        return " | ".join(avis_parts) if avis_parts else "Neutre"

    def refresh_random_symbols(self):
        """Charger 30 symboles ALÉATOIRES depuis popular_symbols."""
        try:
            import random
            all_symbols = self.popular_symbols_data
            if not all_symbols:
                QMessageBox.warning(self, "Erreur", "Aucun symbole populaire disponible")
                return
            
            random_30 = random.sample(all_symbols, min(30, len(all_symbols)))
            
            self.random_list.clear()
            for sym in random_30:
                item = QListWidgetItem(sym)
                item.setData(Qt.UserRole, sym)
                self.random_list.addItem(item)
            
            self.random_label.setText(f"🎲 Aléatoires\n({len(random_30)} symboles)")
        except Exception as e:
            QMessageBox.warning(self, "Erreur", f"Impossible de charger les symboles aléatoires: {e}")
    
    def load_recent_symbols(self):
        """Charger les 30 derniers SYMBOLES AJOUTÉS depuis la base de données."""
        try:
            if SYMBOL_MANAGER_AVAILABLE:
                # Récupérer les 30 derniers symboles directement de la BDD, triés par date d'ajout
                recent = get_recent_symbols(limit=30, active_only=True)
            else:
                # Fallback: si symbol_manager n'est pas disponible, utiliser les listes en fichier
                recent = []
                seen = set()

                for sym in self.mes_symbols_data:
                    if sym not in seen:
                        recent.append(sym)
                        seen.add(sym)

                for sym in reversed(self.popular_symbols_data):
                    if sym not in seen and len(recent) < 30:
                        recent.append(sym)
                        seen.add(sym)

                recent = recent[:30]

            self.recent_list.clear()
            for sym in recent:
                item = QListWidgetItem(sym)
                item.setData(Qt.UserRole, sym)
                self.recent_list.addItem(item)

            self.recent_label.setText(f"🔥 Récents\n({len(recent)} symboles)")
        except Exception as e:
            QMessageBox.warning(self, "Erreur", f"Impossible de charger les symboles récents: {e}")
    
    def _select_all_items(self, list_widget):
        """Sélectionner tous les éléments d'une QListWidget"""
        list_widget.selectAll()

    def _compute_daily_top_movers(self, top_n=30):
        """Récupère les top movers Yahoo (day_gainers/day_losers) via yfinance.screen, sans dépendre des listes locales."""
        now = datetime.now()
        cached = getattr(self, '_daily_movers_cache', None)
        if isinstance(cached, dict):
            ts = cached.get('timestamp')
            if isinstance(ts, datetime):
                if (now - ts).total_seconds() <= 300:
                    return cached

        n = max(1, int(top_n))

        def _extract_entries(screen_payload):
            quotes = []
            if isinstance(screen_payload, dict):
                quotes = screen_payload.get('quotes') or []
            extracted = []
            for q in quotes:
                if not isinstance(q, dict):
                    continue
                symbol = str(q.get('symbol') or '').strip().upper()
                if not symbol:
                    continue
                pct = q.get('regularMarketChangePercent')
                if isinstance(pct, dict):
                    pct = pct.get('raw', pct.get('fmt'))
                try:
                    pct_val = float(pct)
                except Exception:
                    continue
                extracted.append((symbol, pct_val))
            return extracted

        def _fetch_screeners():
            winners_payload = yf.screen('day_gainers', count=n)
            losers_payload = yf.screen('day_losers', count=n)
            winners = _extract_entries(winners_payload)
            losers = _extract_entries(losers_payload)
            return winners, losers

        executor = ThreadPoolExecutor(max_workers=1)
        future = executor.submit(_fetch_screeners)
        try:
            winners, losers = future.result(timeout=12)
        except FuturesTimeoutError:
            raise TimeoutError("Timeout Yahoo Finance: impossible de récupérer les top movers dans le délai imparti.")
        finally:
            executor.shutdown(wait=False, cancel_futures=True)

        cache_payload = {
            'timestamp': now,
            'winners': winners[:n],
            'losers': losers[:n],
        }
        self._daily_movers_cache = cache_payload
        return cache_payload

    def _show_top_daily_movers(self, mover_type):
        """Affiche les top movers du jour et injecte les symboles dans le champ d'analyse."""
        mover_type = (mover_type or '').strip().lower()
        if mover_type not in {'winners', 'losers'}:
            mover_type = 'winners'

        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            top_data = self._compute_daily_top_movers(top_n=30)
        except TimeoutError as e:
            QMessageBox.warning(self, "Timeout", str(e))
            return
        except Exception as e:
            QMessageBox.warning(self, "Erreur", f"Impossible de charger les top movers Yahoo: {e}")
            return
        finally:
            QApplication.restoreOverrideCursor()

        entries = top_data.get(mover_type, []) if isinstance(top_data, dict) else []
        if not entries:
            QMessageBox.information(
                self,
                "Information",
                "Aucune donnée intraday/journalière disponible pour calculer les top movers.",
            )
            return

        symbols = [sym for sym, _pct in entries]
        self.symbol_input.setText(', '.join(symbols))

        title = "Top 30 Winners du jour" if mover_type == 'winners' else "Top 30 Losers du jour"
        lines = []
        for idx, (sym, pct) in enumerate(entries, start=1):
            lines.append(f"{idx:02d}. {sym:8s} {pct:+.2f}%")

        QMessageBox.information(self, title, '\n'.join(lines))

    def _get_clean_columns_and_data(self):
        """Filtrer les colonnes vides ou contenant uniquement des 0"""
        if not self.current_results:
            return [], []
        
        # Récupérer toutes les colonnes
        all_columns = []
        for col in range(self.merged_table.columnCount()):
            header = self.merged_table.horizontalHeaderItem(col)
            if header:
                all_columns.append(header.text())
        
        if not all_columns:
            all_columns = list(self.current_results[0].keys())
        
        # Identifier les colonnes à conserver
        valid_columns = []
        for col in all_columns:
            has_valid_data = False
            for result in self.current_results:
                value = result.get(col, '')
                # Vérifier si la colonne a des données non vides et pas 0
                if value and value != '' and value != 0 and value != '0' and value != 'N/A':
                    try:
                        float_val = float(value) if isinstance(value, str) else value
                        if float_val != 0:
                            has_valid_data = True
                            break
                    except (ValueError, TypeError):
                        if value not in ('', 'N/A', 0, '0'):
                            has_valid_data = True
                            break
            
            if has_valid_data:
                valid_columns.append(col)
        
        print(f"📊 Colonnes filtrées: {len(all_columns)} → {len(valid_columns)} (suppression de {len(all_columns) - len(valid_columns)} colonnes vides/zéro)")
        return valid_columns, self.current_results

    def export_results_csv(self):
        """Exporter les résultats actuels en fichier CSV avec auto-save dans Results"""
        if not hasattr(self, 'current_results') or not self.current_results:
            QMessageBox.warning(self, "Erreur", "Aucun résultat à exporter. Veuillez d'abord lancer une analyse.")
            return
        
        import csv
        from datetime import datetime
        from pathlib import Path
        
        try:
            # Nettoyer les colonnes vides/zéro
            clean_columns, data = self._get_clean_columns_and_data()
            
            if not clean_columns:
                QMessageBox.warning(self, "Erreur", "Aucune colonne avec données valides à exporter")
                return
            
            # Créer le dossier Results avec chemins relatifs
            results_dir = Path(__file__).parent.parent / "Results"
            results_dir.mkdir(parents=True, exist_ok=True)
            
            # Générer le nom du fichier avec timestamp
            filename = f"resultats_analyse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = results_dir / filename
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=clean_columns)
                writer.writeheader()
                
                for result in data:
                    # Créer une ligne en prenant les valeurs existantes
                    row = {}
                    for field in clean_columns:
                        row[field] = result.get(field, '')
                    writer.writerow(row)
            
            QMessageBox.information(self, "Succès", f"✅ Résultats exportés avec succès!\n\nDossier: Results\nFichier: {filename}\nColonnes: {len(clean_columns)}")
            print(f"✅ Résultats exportés en CSV: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export CSV:\n{e}")
            print(f"❌ Erreur export CSV: {e}")
            import traceback
            traceback.print_exc()

    def export_results_excel(self):
        """Exporter les résultats actuels en fichier Excel avec auto-save dans Results"""
        if not hasattr(self, 'current_results') or not self.current_results:
            QMessageBox.warning(self, "Erreur", "Aucun résultat à exporter. Veuillez d'abord lancer une analyse.")
            return
        
        from datetime import datetime
        from pathlib import Path
        
        try:
            # Vérifier si openpyxl est disponible
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                QMessageBox.warning(
                    self,
                    "Bibliothèque manquante",
                    "openpyxl n'est pas installé.\n\nVeuillez installer avec:\npip install openpyxl\n\nOu utiliser l'export CSV à la place."
                )
                return
            
            # Créer le dossier Results avec chemins relatifs
            results_dir = Path(__file__).parent.parent / "Results"
            results_dir.mkdir(parents=True, exist_ok=True)
            
            # Générer le nom du fichier avec timestamp
            filename = f"resultats_analyse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = results_dir / filename
            
            # Nettoyer les colonnes vides/zéro
            columns, data = self._get_clean_columns_and_data()
            
            if not columns:
                QMessageBox.warning(self, "Erreur", "Aucune colonne avec données valides à exporter")
                return
            
            # Créer un classeur
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Résultats"
            
            # En-têtes
            for col_idx, col_name in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = col_name
                # Style d'en-tête
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Données
            for row_idx, result in enumerate(data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = result.get(col_name, '')
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Formatage selon le type
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                    
                    # Coloration selon la valeur pour certaines colonnes
                    if col_name == "Signal":
                        if value == "ACHAT":
                            cell.font = Font(color="00B050")  # Vert
                        elif value == "VENTE":
                            cell.font = Font(color="FF0000")  # Rouge
                    elif col_name in ["Fiabilite", "Rev. Growth (%)", "EBITDA Yield (%)"]:
                        try:
                            val = float(value) if value else 0
                            if val > 0:
                                cell.font = Font(color="00B050")  # Vert
                            elif val < 0:
                                cell.font = Font(color="FF0000")  # Rouge
                        except:
                            pass
                    
                    # Bordures légères
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
            
            # Ajuster les largeurs de colonnes
            for col_idx, col_name in enumerate(columns, 1):
                max_length = max(
                    len(str(col_name)),
                    max(len(str(r.get(col_name, ''))) for r in data) if data else 0
                )
                adjusted_width = min(max_length + 2, 50)  # Max 50 pour lisibilité
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
            
            # Geler la première ligne
            worksheet.freeze_panes = "A2"
            
            # Sauvegarder
            workbook.save(str(file_path))
            
            QMessageBox.information(self, "Succès", f"✅ Résultats exportés avec succès!\n\nDossier: Results\nFichier: {filename}\nColonnes: {len(columns)}")
            print(f"✅ Résultats exportés en Excel: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export Excel:\n{e}")
            print(f"❌ Erreur export Excel: {e}")
            import traceback
            traceback.print_exc()

    def export_results_pdf(self):
        """Exporter tous les graphiques de l'analyse en PDF avec infos du tableau intégrées"""
        try:
            if not hasattr(self, 'current_results') or not self.current_results:
                QMessageBox.warning(self, "Erreur", "Aucun résultat à exporter. Veuillez d'abord lancer une analyse.")
                return
            
            # Construire les colonnes PDF directement depuis les clés des données
            # (éviter _get_clean_columns_and_data qui utilise les headers UI avec \n)
            all_data_keys = []
            for result in self.current_results:
                for key in result.keys():
                    if key not in all_data_keys:
                        all_data_keys.append(key)
            
            # Filtrer : garder uniquement les colonnes qui ont au moins 1 valeur non-vide
            clean_columns = []
            for col in all_data_keys:
                for result in self.current_results:
                    value = result.get(col, '')
                    if value is not None and value != '' and value != 'N/A':
                        try:
                            if isinstance(value, (int, float)) and value != 0:
                                clean_columns.append(col)
                                break
                            elif isinstance(value, str) and value not in ('0', '0.0', '0.00', ''):
                                clean_columns.append(col)
                                break
                        except (ValueError, TypeError):
                            clean_columns.append(col)
                            break
            
            if not clean_columns:
                QMessageBox.warning(self, "Erreur", "Aucune colonne avec données valides à exporter")
                return
            
            print(f"📊 PDF Export: {len(all_data_keys)} clés données → {len(clean_columns)} colonnes retenues")
            
            # Importer le générateur de PDF
            from pdf_generator import PDFReportGenerator
            
            # Créer le générateur et exporter
            generator = PDFReportGenerator()
            min_hold_days = self.min_hold_days_spin.value() if hasattr(self, 'min_hold_days_spin') else 7
            report_meta = {
                'min_holding_days': int(min_hold_days),
            }
            pdf_path = generator.export_pdf(
                self.plots_layout,
                self.current_results,
                clean_columns,
                report_meta=report_meta,
            )
            
            if pdf_path:
                # Extraire juste le nom du fichier pour le message
                from pathlib import Path
                filename = Path(pdf_path).name
                QMessageBox.information(
                    self, 
                    "Succès", 
                    f"✅ PDF créé avec succès!\n\n"
                    f"Dossier: Results\nFichier: {filename}"
                )
            else:
                QMessageBox.critical(self, "Erreur", "Impossible de créer le PDF")
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la création du PDF:\n{e}")
            print(f"❌ Erreur export PDF: {e}")
            import traceback
            traceback.print_exc()


# Ensure the application only launches when run directly
if __name__ == "__main__":
    _install_runtime_diagnostics()
    app = QApplication(sys.argv)
    window = MainWindow()
    window.setWindowTitle("Stock Analysis Tool")
    window.show()
    sys.exit(app.exec_())

    #TODO:
    # - Ajouter un bouton pour exporter les resultats (CSV/Excel)
    # - Ajouter dates d'annonces / résultats dans les signaux (ex: earnings date)
    # - harmoniser l'affichage des plots (embedded + external)
    # - améliorer le threading / gestion des erreurs
    # - Ajouter le earning dates et tous les autres nouveaux criteres a l'analyse et au backtest
    # - Ajouter un bouton pour choisir si backup des resultats avant analyse ou pas