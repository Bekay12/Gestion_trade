"""
ExportMixin — export CSV / Excel / PDF des resultats d'analyse.
"""
from pathlib import Path
from PyQt5.QtWidgets import QMessageBox


class ExportMixin:

    def _get_clean_columns_and_data(self):
        """Filtre les colonnes vides ou contenant uniquement des 0."""
        if not self.current_results:
            return [], []

        all_columns = []
        for col in range(self.merged_table.columnCount()):
            header = self.merged_table.horizontalHeaderItem(col)
            if header:
                all_columns.append(header.text())

        if not all_columns:
            all_columns = list(self.current_results[0].keys())

        valid_columns = []
        for col in all_columns:
            has_valid_data = False
            for result in self.current_results:
                value = result.get(col, '')
                if value and value != '' and value != 0 and value != '0' and value != 'N/A':
                    try:
                        float_val = float(value) if isinstance(value, str) else value
                        if float_val != 0:
                            has_valid_data = True
                            break
                    except (ValueError, TypeError):
                        if value not in ('', 'N/A', 0, '0'):
                            has_valid_data = True
                            break
            if has_valid_data:
                valid_columns.append(col)

        print(f"Colonnes filtrees: {len(all_columns)} -> {len(valid_columns)} "
              f"(suppression de {len(all_columns) - len(valid_columns)} colonnes vides/zero)")
        return valid_columns, self.current_results

    def export_results_csv(self):
        """Exporter les resultats actuels en fichier CSV avec auto-save dans Results."""
        if not self.current_results:
            QMessageBox.warning(self, "Erreur", "Aucun resultat a exporter. Veuillez d'abord lancer une analyse.")
            return

        import csv
        from datetime import datetime

        try:
            clean_columns, data = self._get_clean_columns_and_data()
            if not clean_columns:
                QMessageBox.warning(self, "Erreur", "Aucune colonne avec donnees valides a exporter")
                return

            results_dir = Path(__file__).parent.parent.parent / "Results"
            results_dir.mkdir(parents=True, exist_ok=True)

            filename = f"resultats_analyse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = results_dir / filename

            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=clean_columns)
                writer.writeheader()
                for result in data:
                    row = {field: result.get(field, '') for field in clean_columns}
                    writer.writerow(row)

            self._status(f"✅ Export CSV : Results/{filename} ({len(clean_columns)} colonnes)", 8000)
            print(f"Resultats exportes en CSV: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export CSV:\n{e}")
            import traceback; traceback.print_exc()

    def export_results_excel(self):
        """Exporter les resultats actuels en fichier Excel avec auto-save dans Results."""
        if not self.current_results:
            QMessageBox.warning(self, "Erreur", "Aucun resultat a exporter. Veuillez d'abord lancer une analyse.")
            return

        from datetime import datetime

        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                QMessageBox.warning(
                    self, "Bibliotheque manquante",
                    "openpyxl n'est pas installe.\n\nVeuillez installer avec:\npip install openpyxl\n\nOu utiliser l'export CSV a la place."
                )
                return

            results_dir = Path(__file__).parent.parent.parent / "Results"
            results_dir.mkdir(parents=True, exist_ok=True)

            filename = f"resultats_analyse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = results_dir / filename

            columns, data = self._get_clean_columns_and_data()
            if not columns:
                QMessageBox.warning(self, "Erreur", "Aucune colonne avec donnees valides a exporter")
                return

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Resultats"

            for col_idx, col_name in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_idx, result in enumerate(data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = result.get(col_name, '')
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                    if col_name == "Signal":
                        if value == "ACHAT":
                            cell.font = Font(color="00B050")
                        elif value == "VENTE":
                            cell.font = Font(color="FF0000")
                    elif col_name in ["Fiabilite", "Rev. Growth (%)", "EBITDA Yield (%)"]:
                        try:
                            val = float(value) if value else 0
                            if val > 0:
                                cell.font = Font(color="00B050")
                            elif val < 0:
                                cell.font = Font(color="FF0000")
                        except Exception:
                            pass
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'),  bottom=Side(style='thin')
                    )
                    cell.border = thin_border

            for col_idx, col_name in enumerate(columns, 1):
                max_length = max(
                    len(str(col_name)),
                    max(len(str(r.get(col_name, ''))) for r in data) if data else 0
                )
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

            worksheet.freeze_panes = "A2"
            workbook.save(str(file_path))

            self._status(f"✅ Export Excel : Results/{filename} ({len(columns)} colonnes)", 8000)
            print(f"Resultats exportes en Excel: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export Excel:\n{e}")
            import traceback; traceback.print_exc()

    def export_results_pdf(self):
        """Exporter tous les graphiques de l'analyse en PDF avec infos du tableau integrees."""
        try:
            if not self.current_results:
                QMessageBox.warning(self, "Erreur", "Aucun resultat a exporter. Veuillez d'abord lancer une analyse.")
                return

            all_data_keys = []
            for result in self.current_results:
                for key in result.keys():
                    if key not in all_data_keys:
                        all_data_keys.append(key)

            clean_columns = []
            for col in all_data_keys:
                for result in self.current_results:
                    value = result.get(col, '')
                    if value is not None and value != '' and value != 'N/A':
                        try:
                            if isinstance(value, (int, float)) and value != 0:
                                clean_columns.append(col)
                                break
                            elif isinstance(value, str) and value not in ('0', '0.0', '0.00', ''):
                                clean_columns.append(col)
                                break
                        except (ValueError, TypeError):
                            clean_columns.append(col)
                            break

            if not clean_columns:
                QMessageBox.warning(self, "Erreur", "Aucune colonne avec donnees valides a exporter")
                return

            print(f"PDF Export: {len(all_data_keys)} cles donnees -> {len(clean_columns)} colonnes retenues")

            from pdf_generator import PDFReportGenerator
            generator = PDFReportGenerator()
            min_hold_days = self.min_hold_days_spin.value() if hasattr(self, 'min_hold_days_spin') else 7
            report_meta = {'min_holding_days': int(min_hold_days)}
            pdf_path = generator.export_pdf(
                self.plots_layout,
                self.current_results,
                clean_columns,
                report_meta=report_meta,
            )

            if pdf_path:
                filename = Path(pdf_path).name
                self._status(f"✅ Export PDF : Results/{filename}", 8000)
            else:
                QMessageBox.critical(self, "Erreur", "Impossible de creer le PDF")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la creation du PDF:\n{e}")
            import traceback; traceback.print_exc()
